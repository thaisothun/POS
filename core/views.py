from django.shortcuts import render, redirect
from .models import Inventory, Category, Unit, Customer, PromotionCode, Sale, SaleItem, Store, Purcase, PurchaseItem, Supplier, ExpanseItem, ExpanseCategory, PromotionCodeUsage, InventoryAdjustment, UserProfile
from .forms import InventoryForm, CategoryForm, UnitForm, CustomerForm, InventoryAdjustmentForm, SaleForm, SupplierForm, PurchaseForm, ExpanseForm, ExpanseCategoryForm, PromotionCodeForm, UserForm, UserProfileForm, UserUpdateForm, LoginForm, StoreForm, CustomPasswordChangeForm
from django.http import JsonResponse
import json
from django.db.models import Q
from django.db.models import F
from django.contrib import messages
from django.db.models.signals import post_save
from django.dispatch import receiver
from django.db import transaction
from django.db.models import Value
from django.db.models.functions import Concat
from django.utils import timezone
from datetime import date
from django.db.models import Sum
from datetime import timedelta
from django.db.models import Value
from django.db.models import F
import calendar
from decimal import Decimal
from .function import get_daily_data, get_weekly_data, get_monthly_data, get_top_ten_data, get_top_ten_customer, get_profit, get_trend_daily
from dateutil.relativedelta import relativedelta
from pathlib import Path
from django.conf import settings
from io import BytesIO
from openpyxl import Workbook
from django.http import HttpResponse
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
from django.db.models.functions import TruncDate
from django.db import IntegrityError
from django.contrib.auth.models import User, Group 
from django.contrib.auth import authenticate, login as auth_login
import string
from django.contrib.auth.decorators import login_not_required
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import user_passes_test
from django.core.exceptions import PermissionDenied
from .decorators import in_groups

def index(request):
    user = User.objects.get(username = request.user)
    context = {
        'user' : user,
    }
    return (request, 'index.html', context)

def keep_session(request):

    return JsonResponse({'item' : ''})

@in_groups('manager','superadmin')
def item_management(request):
    items = Inventory.objects.all().filter(status='Active')
    user = User.objects.get(username = request.user)
    if request.method == 'POST':
        form = InventoryForm(request.POST, request.FILES)
        if form.is_valid():           
            form.save()
            messages.success(request,'Item was created sucessfully.')
            return redirect ('item_management')
    else:
        form = InventoryForm()
        form.fields['store'].initial = user.userprofile.store
        form1 = InventoryAdjustmentForm()     
       
    form1 = InventoryAdjustmentForm()
    context = {
        'items' : items,
        'form' : form,
        'form1' : form1,
    }
    
    return render(request, 'item_management.html', context)

@in_groups('manager','superadmin')
def item_management_deactivate(request):
    items = Inventory.objects.all().filter(status='Inactive')
    if request.method == 'POST':
        form = InventoryForm(request.POST, request.FILES)
        if form.is_valid():            
            form.save()
            messages.success(request,'Item was deactivated sucessfully.')
            return redirect ('item_management')
    else:
        form = InventoryForm()     
    
    context = {
        'items' : items,
        'form' : form,
    }
    
    return render(request, 'item_management_deactivate.html', context)

@in_groups('manager','superadmin')
def sort_inventory(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(status='Active').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def sort_inventory_deactivate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(status='Inactive').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def search_inventory(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Active').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def search_inventory_deactivate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Inactive').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def deactivate_inventory(request, product_id):
    if request.method == 'POST':
        item = Inventory.objects.get(id=product_id)
        if item.quantity == 0:
            item.status = 'Inactive'
            item.save()
            message = f'Item ID No. {product_id} was successfully deactivated.'
            return JsonResponse({'message_sucess': message}, safe=False)
        else:
            message = f'Item ID No. {product_id} remains {item.quantity} {item.unit}. Please make sure the quantity is zeroized before perform the action.'
            return JsonResponse({'message_error': message}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def activate_inventory(request, product_id):
    if request.method == 'POST':
        try:    
            item = Inventory.objects.get(id=product_id)      
            item.status = 'Active'
            item.save()
            message = f'Item ID No. {product_id} was successfully activated.'
            return JsonResponse({'message_sucess': message}, safe=False)
        except Exception as e:
            return JsonResponse({'message_error': str(e)}, safe=False)
        except IntegrityError as e:
            return JsonResponse({'message_error': str(e)}, safe=False)
    return JsonResponse({"message_error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def get_inventory(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('product_id')
        request.session['temp_product_id'] = product_id 
        item = list(Inventory.objects.values('id','name','store','category','size','price','quantity','barcode_number','description','photo','unit','cost', 'reorder_alert','status').filter(id=product_id))   
    return JsonResponse(item, safe=False)

@in_groups('manager','superadmin')
def save_inventory(request, product_id):
    if request.method == 'POST':
        update_item = Inventory.objects.get(id = product_id)
        form = InventoryForm(request.POST, request.FILES, instance=update_item)
        if form.is_valid():
            try:
                form.save()
                return JsonResponse({"message_sucess": "Item was updated sucessfully."})
            except Exception as e:
                return JsonResponse({"message_error": str(e)})
            except IntegrityError as e:
                return JsonResponse({"message_error": str(e)})
        return JsonResponse({"message_error": 'form error'})       
    return JsonResponse({"error": "Invalid request"})

@in_groups('manager','superadmin')
def reminder_inventory(request):
    items = Inventory.objects.all().filter(status='Active').filter(quantity__lt=F('reorder_alert'))
    context = {
        'items' : items
    }
    return render(request, 'reminder_inventory.html', context)

@in_groups('manager','superadmin')
def reminder_inventory_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(status='Active').filter(quantity__lt=F('reorder_alert')).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def reminder_inventory_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Active').filter(quantity__lt=F('reorder_alert')).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def category(request):
    items_category = Category.objects.all()
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():            
            with transaction.atomic():
                try:    
                    form.save()
                    messages.success(request, 'Category was created sucessfully.')
                    return redirect ('category')
                except Exception as e:
                    transaction.set_rollback(True)
                    messages.error(request, str(e))
                except IntegrityError as e:
                    transaction.set_rollback(True)
                    messages.error(request, str(e))
    else:
        form = CategoryForm() 
    
    context = {
        'items' : items_category,
        'form' : form,
    }

    return render(request, 'category.html', context)

@in_groups('manager','superadmin')
def category_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        items = list(Category.objects.values('id','name','top_ten_filter').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def category_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Category.objects.values('id','name', 'top_ten_filter').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def get_category(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('product_id')
        request.session['temp_product_id'] = product_id 
        item = list(Category.objects.values('id','name','top_ten_filter').filter(id=product_id))   
    return JsonResponse(item, safe=False)

@in_groups('manager','superadmin')
def category_modify(request, product_id):
    if request.method == 'POST':
        update_category = Category.objects.get(name = product_id)
        form = CategoryForm(request.POST, instance=update_category)
        if form.is_valid():
            try:
                form.save()
                return JsonResponse({"message_sucess": "Category was updated sucessfully."})
            except Exception as e:
                return JsonResponse({"message_error": str(e)})  
            except IntegrityError as e:
                return JsonResponse({"message_error": str(e)})          
        return JsonResponse(form.errors)       
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def category_delete(request, product_id):
    if request.method == 'POST':
        if (Inventory.objects.filter(category__name = product_id).exists()):
            return JsonResponse({"message_error": "There are items under this category. Please make sure all items under this category have been removed before perform the action."})
        else:
            delete_category = Category.objects.get(name = product_id)
            delete_category.delete()
            return JsonResponse({"message_sucess": "Category was deleted sucessfully."})            
    return JsonResponse({"error": "Invalid request"})

@in_groups('manager','superadmin')
def unit(request):
    items_unit = Unit.objects.all()
    if request.method == 'POST':
        form = UnitForm(request.POST)
        if form.is_valid():            
            with transaction.atomic():    
                try:
                    form.save()
                    messages.success(request, 'Unit was created sucessfully.')
                    return redirect ('unit')
                except Exception as e:
                    transaction.set_rollback(True)
                    messages.error(request, str(e))
                except IntegrityError as e:
                    transaction.set_rollback(True)
                    messages.error(request, str(e))
    else:
        form = UnitForm() 
    
    context = {
        'items' : items_unit,
        'form' : form,
    }

    return render(request, 'unit.html', context)

@in_groups('manager','superadmin')
def unit_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Unit.objects.values('id','name').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def unit_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        items = list(Unit.objects.values('id','name').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def get_unit(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('product_id')
        request.session['temp_product_id'] = product_id 
        item = list(Unit.objects.values('id','name').filter(id=product_id))
    return JsonResponse(item, safe=False)

@in_groups('manager','superadmin')
def unit_modify(request, product_id):
    if request.method == 'POST':
        update_category = Unit.objects.get(name = product_id)
        form = UnitForm(request.POST, instance=update_category)
        if form.is_valid():
            try:    
                form.save()
                return JsonResponse({"message_sucess": "Unit was updated sucessfully."})
            except Exception as e:
                return JsonResponse({"message_error": str(e)})  
            except IntegrityError as e:
                return JsonResponse({"message_error": str(e)})          
        return JsonResponse(form.errors)              
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def unit_delete(request, product_id):
    if request.method == 'POST':
        if (Inventory.objects.filter(unit__name = product_id).exists()):
            return JsonResponse({"message_error": "There are still items under this unit."})
        else:
            delete_category = Unit.objects.get(name = product_id)
            delete_category.delete()
            return JsonResponse({"message_sucess": "Category was deleted sucessfully."})            
    return JsonResponse({"error": "Invalid request"})

@in_groups('manager','superadmin')
def category_filter(request, name):
    if Inventory.objects.all().filter(status='Active').filter(category__name=name).exists():
        items = Inventory.objects.all().filter(status='Active').filter(category__name=name)
        category=None     
    else:
        category = name
        items = None
    form = InventoryForm()
    form1 = InventoryAdjustmentForm()
    context = {
        'items' : items,
        'form' : form,
        'category' : category,
        'form1' : form1 
    }
    
    return render(request, 'category_filter.html', context)

@in_groups('manager','superadmin')
def category_filter_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        category = data.get('category')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(status='Active').filter(category__name=category).order_by(f'{sort_method}{sortby}'))
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def category_filter_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        category = data.get('category')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Active').filter(category__name=category).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def unit_filter(request, name):
    if Inventory.objects.all().filter(status='Active').filter(unit__name=name).exists():
        items = Inventory.objects.all().filter(status='Active').filter(unit__name=name)
        unit=None     
    else:
        unit = name
        items = None
    form = InventoryForm()
    context = {
        'items' : items,
        'form' : form,
        'unit' : unit 
    }
    
    return render(request, 'unit_filter.html', context)

@in_groups('manager','superadmin')
def unit_filter_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        unit = data.get('unit')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(status='Active').filter(unit__name=unit).order_by(f'{sort_method}{sortby}'))
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def unit_filter_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        unit = data.get('unit')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Active').filter(unit__name=unit).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def print_barcode(request):
    items = Inventory.objects.all().filter(status='Active')
    context= {
        'items' : items
    }
    return render(request, 'print_barcode.html', context)

@in_groups('manager','superadmin')
def sort_inventory_print_barcode(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(status='Active').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def search_inventory_print_barcode(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Active').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def get_item_print_barcode(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('product_id')
        request.session['temp_product_id'] = product_id 
        item = list(Inventory.objects.values('id','name','store','category','size','price','quantity','barcode_image','description','photo','unit','cost', 'reorder_alert').filter(id=product_id))   
    return JsonResponse(item, safe=False)

def customer_management(request):
    items = Customer.objects.all().filter(status='Active')
    user = User.objects.get(username = request.user)
    if request.method == 'POST':
        form = CustomerForm(request.POST, request.FILES)
        if form.is_valid():            
            try:
                form.save()
                messages.success(request,'Customer was created sucessfully.')
                return redirect ('customer_management')
            except Exception as e:
                messages.error(request, e )
                return redirect ('customer_management')
            except IntegrityError as e:
                messages.error(request, e )
                return redirect ('customer_management')
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '
        messages.error(request, error)    
    
    else:
        form = CustomerForm()
        form.fields['store'].initial = user.userprofile.store
    
    context = {
        'items' : items,
        'form' : form,
    }
    return render(request, 'customer_management.html', context)

def customer_management_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        items = list(Customer.objects.values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','note','created_on').filter(status='Active').order_by(f'{sort_method}{sortby}'))

    return JsonResponse(items, safe=False)

def customer_management_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Customer.objects.values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','note','created_on').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(Q(full_name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Active').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def deactivate_customer(request, product_id):
    if request.method == 'POST':
        item = Customer.objects.get(id=product_id)
        try:
            customer_balance = item.credit_balance.all().order_by('created_at').last().balance
        except:
            customer_balance = 0
        if customer_balance == 0:
            item.status = 'Inactive'
            item.save()
            message = f'Customer with ID No. {product_id} was successfully deactivated.'
            return JsonResponse({'message_sucess': message}, safe=False)
        else:
            message = f'customer with Id No. {product_id} remains {customer_balance} credit balance. Please make sure the credit balance of customer is zeroized before perform the action.'
            return JsonResponse({'message_error': message}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

def get_customer(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('product_id')
        request.session['temp_product_id'] = product_id 
        item = list(Customer.objects.values('id','store__id','first_name','last_name','gender','phone_number1','e_mail','membership','note','created_on','address','status').filter(id=product_id))  
         
    return JsonResponse(item, safe=False)

def save_customer(request, product_id):
    if request.method == 'POST':
        update_item = Customer.objects.get(id = product_id)
        form = CustomerForm(request.POST, request.FILES, instance=update_item)
        if form.is_valid():
            try:
                form.save()
                return JsonResponse({"message_sucess": "Customer was updated sucessfully."})
            except Exception as e:
                return JsonResponse({"message_error": str(e)})
            except IntegrityError as e:
                return JsonResponse({"message_error": str(e)})
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '
        return JsonResponse({"message_error": str(error)})           
    return JsonResponse({"error": "Invalid request"})

@in_groups('manager','superadmin')
def deactivate_customer_list(request):
    items = Customer.objects.all().filter(status='Inactive')
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():            
            form.save()
            messages.success(request,'customer was created sucessfully.')
            return redirect ('customer_management')
    else:
        form = CustomerForm()     
    
    context = {
        'items' : items,
        'form' : form,
    }
    
    return render(request, 'customer_management_deactivate.html', context)

@in_groups('manager','superadmin')
def activate_customer(request, product_id):
    if request.method == 'POST':
        try:    
            item = Customer.objects.get(id=product_id)      
            item.status = 'Active'
            item.save()
            message = f'Customer ID No. {product_id} was successfully activated.'
            return JsonResponse({'message_sucess': message}, safe=False)
        except Exception as e:
            return JsonResponse({'message_error': str(e)}, safe=False)
        except IntegrityError as e:
            return JsonResponse({'message_error': str(e)}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def customer_management_deactivate_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        items = list(Customer.objects.values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','note','created_on').filter(status='Inactive').order_by(f'{sort_method}{sortby}'))

    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def customer_management_deactivate_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Customer.objects.values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','note','created_on').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(Q(full_name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Inactive').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def get_inventory_adjustment(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('product_id')
        request.session['temp_product_id'] = product_id 
        item = list(Inventory.objects.values('id','name','store','category','size','price','quantity','barcode_number','description','photo','unit','cost', 'reorder_alert').filter(id=product_id))   
    return JsonResponse(item, safe=False)

@in_groups('manager','superadmin')
def save_inventory_adjustment(request):
    if request.method == 'POST':
        form = InventoryAdjustmentForm(request.POST)
        user = User.objects.get(username = request.user)
        if form.is_valid():
            with transaction.atomic():
                try:
                    adjustment = form.save(commit=False)
                    adjustment.store = Store.objects.get(name=user.userprofile.store) 
                    adjustment.save()            
                    return JsonResponse({"message_sucess": "Adjustment was updated sucessfully."})
                except Exception as e:
                    transaction.set_rollback(True)
                    return JsonResponse({"message_error": str(e)})
                except IntegrityError as e:
                    transaction.set_rollback(True)
                    return JsonResponse({"message_error": str(e)})
        return JsonResponse({"message_error": 'form errors'})
    return JsonResponse({"error": "Invalid request"})

@receiver(post_save, sender=Sale)
def sale_transaction_save_item(sender, instance, created, **kwargs):
    if created:
        request = getattr(instance, '_request', None)
        try:
            with transaction.atomic():
                cart_item = json.loads(instance.cart_item)
                user= User.objects.get(username = request.user)
                promotion_discount = PromotionCode.objects.get(name=instance.promotion_code).amount_discount
                if instance.customer:
                    customer =  Customer.objects.get(id=instance.customer.id)
                    if PromotionCode.is_valid(instance.promotion_code, customer) == False:
                        raise Exception('Promotion Code Limit Excess')
                    else:
                        PromotionCodeUsage.objects.create(promotion=PromotionCode.objects.get(name=instance.promotion_code), customer= Customer.objects.get(id=customer.id), sale_id=Sale.objects.get(id=instance.id), discount_amount = promotion_discount)
                else: 
                    PromotionCodeUsage.objects.create(promotion=PromotionCode.objects.get(name=instance.promotion_code), customer= None, sale_id=Sale.objects.get(id=instance.id), discount_amount = promotion_discount)
                
                for item in cart_item:
                    item['sale_id'] = instance.id
                    item['unit_price'] = float(item['unit_price'])
                    del item['photo']
                    del item['name']
                for data in cart_item:
                    sale_item = SaleItem(**data)
                    sale_item.save()
                    InventoryAdjustment.objects.create(name=Inventory.objects.get(id=data['item_id']), adjustment_type = 'Out', type='Sale', quantity = data['quantity'], description = f'add sale item {data['item_id']} sale id {instance.id}',store= Store.objects.get(name=user.userprofile.store))
                    messages.success(request, f'Sale invoice has been recorded successfully.')
        except Exception as e:
            delete_sale = Sale.objects.get(id=instance.id)
            delete_sale.delete()
            messages.error(request, f'Sale invoice recorded fail. {e}')
            return redirect('sale_transaction')

@in_groups('cashier')
def sale_transaction(request):
    category_item = Category.objects.all()
    items = Inventory.objects.all().filter(status='Active')
    promotion_code_def = PromotionCode.objects.get(name='no code')
    user= User.objects.get(username = request.user)
    active_promotion_code = []
    for item in PromotionCode.objects.all():
        if item.is_valid() == True:
            active_promotion_code.append(item.name)
    if request.method == 'POST':
        form = SaleForm(request.POST)
        cart_item = request.POST.get('cart_save_item')
        if form.is_valid():
            try:
                item = form.save(commit=False)
                item.cart_item = cart_item
                item._request = request
                item.user = request.user
                item.save()  
                return redirect('sale_transaction')
            except Exception as e:
                messages.error(request,f'Sale invoice recorded fail. {e}')
            except IntegrityError as e:
                messages.error(request,f'Sale invoice recorded fail. {e}')
    else:
        form = SaleForm()
        form.fields['promotion_code'].queryset  = PromotionCode.objects.filter(name__in = active_promotion_code)
        form.fields['promotion_code'].initial = promotion_code_def
        form.fields['store'].initial = Store.objects.get(name = user.userprofile.store)
        
    form = SaleForm()
    form.fields['promotion_code'].queryset  = PromotionCode.objects.filter(name__in = active_promotion_code)
    form.fields['promotion_code'].initial = promotion_code_def
    form.fields['store'].initial = Store.objects.get(name = user.userprofile.store)
    form1 = CustomerForm()
    form1.fields['store'].initial = Store.objects.get(name = user.userprofile.store)
    
    context = {
        'category_item' : category_item,
        'items' : items,
        'form' : form,
        'form1' : form1
    }

    return render(request, 'sale_transaction.html', context)

def sale_transaction_get_category(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        item_category = data.get('category').lower()
        if item_category == 'all items':
            items = list(Inventory.objects.values('id','name','store','category','size','price','quantity','barcode_number','description','photo','unit','cost', 'reorder_alert').filter(status='Active'))  
        else:
            items = list(Inventory.objects.values('id','name','store','category','size','price','quantity','barcode_number','description','photo','unit','cost', 'reorder_alert').filter(status='Active').filter(category__name=item_category))   
        
    return JsonResponse(items, safe=False)

def sale_transaction_get_promotion_code(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id').lower()
        items = list(PromotionCode.objects.values('id','amount_discount').filter(id=id))  
        
    return JsonResponse(items, safe=False)

def sale_transaction_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        item_category = data.get('category').lower()
        if item_category == 'all items':
            items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert','photo','cost').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Active')) 
        else:       
            items = list(Inventory.objects.values('id','name','store__name','category__name','size','price','quantity','reorder_alert','photo', 'cost').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Active').filter(category__name=item_category))
    return JsonResponse(items, safe=False)

def sale_transaction_print_receipt(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        if (id):
            sale_id = list(Sale.objects.values('id','store__name','store__address','store__logo','invoice_number','transaction_date','customer__first_name','customer__last_name','customer__address','customer__phone_number1','total_amount','discount','promotion_code__amount_discount','net_amount','payment_method','instruction').filter(id=id))
            sale_item = list(SaleItem.objects.values('item__name','quantity','unit_price','total_price').filter(sale=id))
        else:
            last_record = Sale.objects.last()
            sale_id = list(Sale.objects.values('id','store__name','store__address','store__logo','invoice_number','transaction_date','customer__first_name','customer__last_name','customer__address','customer__phone_number1','total_amount','discount','promotion_code__amount_discount','net_amount','payment_method','instruction').filter(id=last_record.id))
            sale_item = list(SaleItem.objects.values('item__name','quantity','unit_price','total_price').filter(sale=last_record.id))
    
    return JsonResponse(data={'sale_id' : sale_id, 'sale_item' : sale_item}, safe=False) 

def sale_transaction_search_customer(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        customer_list = list(Customer.objects.values('first_name','last_name','id', 'phone_number1','phone_number2').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(Q(full_name__icontains=search_text) | Q(phone_number1__icontains=search_text) | Q(phone_number2__icontains=search_text)|Q(id__icontains=search_text)))
    return JsonResponse(customer_list, safe=False)

def sale_transaction_add_customer(request):
    if request.method == 'POST':
        form = CustomerForm(request.POST)
        try:    
            if form.is_valid():
                form.save()
                last_customer = Customer.objects.last()
                customer = list(Customer.objects.values('first_name','last_name','id').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(id=last_customer.id))
                return JsonResponse({"message_sucess": "Customer was added sucessfully.", 'customer' : customer })    
        except Exception as e:
            return JsonResponse({"message_error": str(e) })
        except IntegrityError as e:
            return JsonResponse({"message_error": str(e) })
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '            
        return JsonResponse({"form_error": error })      
    return JsonResponse({"error": "Invalid request"})

def sale_transaction_list(request):
    now = timezone.now()
    category_item = Category.objects.all()
    items = Inventory.objects.all().filter(status='Active')
    sale_transactions = Sale.objects.all().filter(transaction_date__month = now.month).filter(deleted_at__isnull=True).order_by('-invoice_number')
    form = SaleForm()
    context = {
        'sale_transactions' : sale_transactions,
        'form' : form,
        'category_item' : category_item,
        'items' : items
    }
    return render(request, 'sale_transaction_list.html', context)

def sale_transaction_list_filter_date(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        sale_transactions = list(Sale.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)))
    return JsonResponse(sale_transactions, safe=False)        
    
def sale_transaction_list_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        
        items = list(Sale.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status', 'id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False) 

def sale_transaction_list_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Sale.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(Q(invoice_number__icontains=search_text) | Q(full_name__icontains=search_text)).order_by(f'{sort_method}{sortby}'))
    return JsonResponse(items, safe=False)

def sale_transaction_list_delete_submit(request, id):
    if request.method == 'POST':
        sale_transaction = Sale.objects.get(invoice_number=id)
        file_path = Path(settings.BASE_DIR / 'core/static/data')/'sale_pending_delete.json'
        if file_path.exists():
            with open(file_path, 'r') as file:
                data = json.load(file)
        else: 
            data = {}
        
        data[sale_transaction.id] = ({
            'sale_id' : sale_transaction.id,
            'status' : sale_transaction.status,
            'payment_method' : sale_transaction.payment_method
        })
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=4)

        if sale_transaction.status == 'Pending Delete':
            message = f'Invoice No. {id} was already sumitted for approval.'
            return JsonResponse({'message_error': message}, safe=False)
        else:    
            sale_transaction.payment_method = 'Cash'
            sale_transaction.status = 'Pending Delete'
            sale_transaction.save()
            message = f'Invoice No. {id} was successfully sumitted for approval.'
            return JsonResponse({'message_sucess': message}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

def sale_transaction_list_filter_payment(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        payment_method = data.get('payment_method')
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        items = list(Sale.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status', 'id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(payment_method = payment_method))
    return JsonResponse(items, safe=False)

def sale_transaction_list_filter_date(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        sale_transactions = list(Sale.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)))
    return JsonResponse(sale_transactions, safe=False)      

def get_sale_transaction(request, id):
    if request.method == 'POST':
        sale_transaction = list(Sale.objects.values('store__name','invoice_number','transaction_date','user','customer__first_name','customer__last_name','total_amount','discount','promotion_code','net_amount','payment_method','status', 'id','promotion_code__amount_discount').filter(id=id))
        sale_item = list(SaleItem.objects.values('item__photo','item__name','quantity','unit_price','item__id').filter(sale=id))
        for item in sale_item:
            if item['item__photo']:
                item['item__photo_url'] = f"{settings.MEDIA_URL}media/{item['item__photo']}"
            else:
                 item['item__photo_url'] = None   
        items = {
            'sale_transaction' : sale_transaction,
            'sale_item' : sale_item
        }
    return JsonResponse(items, safe=False) 

def sale_transaction_save(request):
    if request.method == "POST":
        data = json.loads(request.body)
        cart = data.get('cart')
        user = User.objects.get(username=request.user)
        invoice_number_id = data.get('invoice_number_id')
        invoice_number = data.get('invoice_number')
        sale_ids = [item['item_id'] for item in cart]
        print(sale_ids)
        with transaction.atomic():     
            try:
                remove_sale = SaleItem.objects.filter(sale=invoice_number_id).exclude(item__in=sale_ids)
                for item in remove_sale:
                    InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item.item_id), adjustment_type = 'In', type='Adjustment', quantity = item.quantity, store=Store.objects.get(name=user.userprofile.store), description = f'remove sale item {item.item_id} sale id {item.sale}')
                    item.delete()
            except Exception as e:
                transaction.set_rollback(True)
                return JsonResponse({'message_error': str(e)}, safe=False)
            for item in cart:
                if SaleItem.objects.filter(sale=invoice_number_id).filter(item=item['item_id']).exists():
                    try:
                        update_sale_item = SaleItem.objects.filter(sale=invoice_number_id).get(item=item['item_id'])
                        different_item_quantiry = item['quantity'] - update_sale_item.quantity
                        if different_item_quantiry > 0:
                            InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'Out', type='Sale', quantity = different_item_quantiry, store=Store.objects.get(name=user.userprofile.store), description = f'increase quantity sale item {item['item_id']} sale id {invoice_number_id}')
                        if different_item_quantiry < 0:
                            InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'In', type='Adjustment', quantity = abs(different_item_quantiry), store=Store.objects.get(name=user.userprofile.store), description = f'decrease quantity sale item {item['item_id']} sale id {invoice_number_id}')
                        update_sale_item.quantity = item['quantity']
                        update_sale_item.unit_price = float(item['unit_price'])
                        update_sale_item.save()
                    except Exception as e:
                        transaction.set_rollback(True)
                        return JsonResponse({'message_error': str(e)}, safe=False)
                else:
                    try:
                        sale=Sale.objects.get(id=invoice_number_id)
                        new_sale_item = Inventory.objects.get(id=item['item_id'])
                        SaleItem.objects.create(sale=sale, item=new_sale_item, quantity= item['quantity'], unit_price=float(item['unit_price']), cost=new_sale_item.cost)
                        InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'Out', type='Sale', quantity = item['quantity'], store=Store.objects.get(name=user.userprofile.store), description = f'add sale item {item['item_id']} sale id {invoice_number_id}')                
                    except Exception as e:
                        transaction.set_rollback(True)
                        return JsonResponse({'message_error': str(e)}, safe=False)
        try:
            sale_item = Sale.objects.get(id=invoice_number_id)
            sale_item.discount = data.get('discount')
            sale_item.promotion_code = PromotionCode.objects.get(id= data.get('promotion_code'))
            sale_item.payment_method = data.get('payment_method')
            sale_item.save()
            message = f'Invoice No. {invoice_number} was successfully saved.'
            return JsonResponse({'message_sucess': message }, safe=False)
        except Exception as e:
            return JsonResponse({'message_error': str(e)}, safe=False)

def sale_transaction_paid(request):
    if request.method == "POST":
        data = json.loads(request.body)
        user = User.objects.get(username=request.user)
        cart = data.get('cart')
        invoice_number_id = data.get('invoice_number_id')
        invoice_number = data.get('invoice_number')
        sale_ids = [item['item_id'] for item in cart]
        with transaction.atomic():     
            try:
                remove_sale = SaleItem.objects.filter(sale=invoice_number_id).exclude(item__in=sale_ids)
                for item in remove_sale:
                    InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item.item_id), adjustment_type = 'In', type='Adjustment', quantity = item.quantity, store=Store.objects.get(name=user.userprofile.store), description = f'remove sale item {item.item_id} sale id {item.sale}')
                    item.delete()
            except Exception:
                transaction.set_rollback(True)
                return JsonResponse({'message_error': str(e)}, safe=False)
            for item in cart:
                if SaleItem.objects.filter(sale=invoice_number_id).filter(item=item['item_id']).exists():
                    try:
                        update_sale_item = SaleItem.objects.filter(sale=invoice_number_id).get(item=item['item_id'])
                        different_item_quantiry = item['quantity'] - update_sale_item.quantity
                        if different_item_quantiry > 0:
                            InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'Out', type='Sale', quantity = different_item_quantiry, store=Store.objects.get(name=user.userprofile.store), description = f'increase quantity sale item {item['item_id']} sale id {invoice_number_id}')
                        if different_item_quantiry < 0:
                            InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'In', type='Adjustment', quantity = abs(different_item_quantiry), store=Store.objects.get(name=user.userprofile.store), description = f'decrease quantity sale item {item['item_id']} sale id {invoice_number_id}')
                        update_sale_item.quantity = item['quantity']
                        update_sale_item.unit_price = float(item['unit_price'])
                        update_sale_item.save()
                    except Exception as e:
                        transaction.set_rollback(True)
                        return JsonResponse({'message_error': str(e)}, safe=False)
                else:
                    try:
                        sale=Sale.objects.get(id=invoice_number_id)
                        new_sale_item = Inventory.objects.get(id=item['item_id'])
                        SaleItem.objects.create(sale=sale, item=new_sale_item, quantity= item['quantity'], unit_price=float(item['unit_price']), cost=new_sale_item.cost)
                        InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'Out', type='Sale', quantity = item['quantity'], store=Store.objects.get(name=user.userprofile.store), description = f'add sale item {item['item_id']} sale id {invoice_number_id}')                
                    except Exception as e:
                        transaction.set_rollback(True)
                        return JsonResponse({'message_error': str(e)}, safe=False)
        try:
            sale_transaction = Sale.objects.get(id=invoice_number_id)
            sale_transaction.promotion_code = PromotionCode.objects.get(id=data.get('promotion_code'))
            sale_transaction.status = 'Paid'
            sale_transaction.discount = data.get('discount')
            sale_transaction.payment_method = data.get('payment_method')
            sale_transaction.save()
            message = f'Invoice No. {invoice_number} was successfully paid.'
            return JsonResponse({'message_sucess': message}, safe=False) 
        except Exception as e:
            return JsonResponse({'message_error': str(e)}, safe=False)

@in_groups('manager','superadmin')
def sale_transaction_list_approval(request):
    sale_transactions = Sale.objects.all().filter(deleted_at__isnull=True).filter(status="Pending Delete").order_by('-invoice_number')
    context = {
        'sale_transactions' : sale_transactions
    }

    return render(request, 'sale_transaction_list_approval.html', context)

@in_groups('manager','superadmin')
def sale_transaction_list_filter_date_approval(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        sale_transactions = list(Sale.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(status="Pending Delete"))
    return JsonResponse(sale_transactions, safe=False)    

@in_groups('manager','superadmin')
def sale_transaction_list_sort_approval(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        
        items = list(Sale.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status', 'id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(status="Pending Delete").order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def sale_transaction_list_search_approval(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Sale.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(status="Pending Delete").filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(Q(invoice_number__icontains=search_text) | Q(full_name__icontains=search_text)).order_by(f'{sort_method}{sortby}'))
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def sale_transaction_list_delete_approval(request, id):
    if request.method == 'POST':
        sale_transaction = Sale.objects.get(invoice_number=id)
        sale_transaction.delete_custom()
        message = f'Invoice No. {id} was successfully deleted.'
        return JsonResponse({'message_sucess': message}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def sale_transaction_list_delete_reject(request, id):
    if request.method == 'POST':
        sale_transaction = Sale.objects.get(invoice_number=id)
        file_path = Path(settings.BASE_DIR /'core/static/data')/'sale_pending_delete.json'
        with open(file_path, 'r') as file:
            data = json.load(file)
        if data[sale_transaction.id]['sale_id'] == sale_transaction.id:
            sale_transaction.status = data[sale_transaction.id]['status']
            sale_transaction.payment_method = data[sale_transaction.id]['payment_method']            
            del data[sale_transaction.id]
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=4)
        sale_transaction.save()        
        message = f'Invoice No. {id} was successfully rejected.'
        return JsonResponse({'message_sucess': message}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def sale_transaction_list_promotion_code(request):
    items_category = PromotionCode.objects.all()
    if request.method == 'POST':
        form = PromotionCodeForm(request.POST)
        if form.is_valid():            
            form.save()
            messages.success(request, 'Promotion Code was created sucessfully.')
            return redirect ('sale_transaction_list_promotion_code')      
    else:
        form = PromotionCodeForm()
        user = User.objects.get(username = request.user)
        form.fields['store'].initial = user.userprofile.store 
    
    context = {
        'items' : items_category,
        'form' : form,
    }

    return render(request, 'sale_transaction_promotion_code.html', context)

@in_groups('manager','superadmin')
def sale_transaction_list_promotion_code_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        items = list(PromotionCode.objects.values('id','name', 'amount_discount','max_uses','max_uses_per_user','start_date','end_date','store__name').order_by(f'{sort_method}{sortby}'))
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def sale_transaction_list_promotion_code_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(PromotionCode.objects.values('id','name', 'amount_discount','max_uses','max_uses_per_user','start_date','end_date').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def sale_transaction_list_promotion_code_get(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('product_id')
        request.session['temp_product_id'] = product_id 
        item = list(PromotionCode.objects.values('id','name', 'amount_discount','max_uses','max_uses_per_user','start_date','end_date','store').filter(id=product_id))   
    return JsonResponse(item, safe=False)

@in_groups('manager','superadmin')
def sale_transaction_list_promotion_code_modify(request, product_id):
    if request.method == 'POST':
        update_category = PromotionCode.objects.get(id = product_id)
        form = PromotionCodeForm(request.POST, instance=update_category)
        if form.is_valid():
            try:
                form.save()
                return JsonResponse({"message_sucess": "Promotion code was updated sucessfully."})
            except Exception as e:
                return JsonResponse({"message_error": str(e)})
            except IntegrityError as e:
                return JsonResponse({"message_error": str(e)})
        return JsonResponse({"error": form.errors})       
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','cashier')
def sale_transaction_list_promotion_code_delete(request, product_id):
    if request.method == 'POST':
        if (Sale.objects.filter(promotion_code__name = product_id).exists()):
            return JsonResponse({"message_error": "Promotion code already used and can not be deleted."})
        else:
            delete_category = PromotionCode.objects.get(name = product_id)
            delete_category.delete()
            return JsonResponse({"message_sucess": "Promotion code was deleted sucessfully."})            
    return JsonResponse({"error": "Invalid request"})

@in_groups('manager','superadmin')
def sale_transaction_list_promotion_code_filter(request, name):
    if PromotionCode.objects.all().filter(name=name).exists():
        items = Sale.objects.all().filter(promotion_code__name=name)
        category=None     
    else:
        category = name
        items = None
    form = PromotionCodeForm()
    context = {
        'items' : items,
        'form' : form,
        'category' : category, 
    }
    
    return render(request, 'sale_transaction_promotion_code_filter.html', context)

def sale_transaction_list_statistics(request):
    total_sale_month = Sale.objects.filter(status='Paid').filter(transaction_date__date__range=(timezone.now().replace(day=1),timezone.now())).aggregate(total=Sum('total_amount'))['total'] or 0
    total_sale_pending = Sale.objects.filter(status='Pending').aggregate(total=Sum('total_amount'))['total'] or 0
    pending_count = Sale.objects.filter(status='Pending').filter(deleted_at__isnull=True).count()
    pending_paid_3days_ago = Sale.objects.filter(status='Pending').filter(transaction_date__date__lte=timezone.now() - timedelta(days=3)).aggregate(total=Sum('total_amount'))['total'] or 0    
    sale_target = Store.objects.get(id=1)
    days_month = calendar.monthrange(timezone.now().year, timezone.now().month)
    pro_target = (sale_target.sale_target/days_month[1])*int(timezone.now().day) 
    achieve = (Decimal(total_sale_month) / (pro_target))*100
    short_amount = Decimal(sale_target.sale_target) - total_sale_month
    
    data_sale_trend, total = get_trend_daily(7,Sale)
    sale_trend = []
    for item in data_sale_trend:
        sale_trend.append(item['total'])
    
    data_sale_day_current_month, total_sale_current_month = get_daily_data(timezone.now().replace(day=1),timezone.now(), Sale)
    sale_day_current_month = []
    labels = []
    for item in data_sale_day_current_month:
        sale_day_current_month.append(item['total'])
        labels.append(item['day'])

    data_sale_day_last_month, total_sale_last_month = get_daily_data((timezone.now()-relativedelta(months=1)).replace(day=1), timezone.now() - relativedelta(months=1),Sale)
    sale_day_last_month = []
    for item in data_sale_day_last_month:
        sale_day_last_month.append(item['total'])
    
    if total_sale_current_month['total'] == 0:
        sale_performance = 0
        message = ''
    else:
        sale_performance = (total_sale_current_month['total'] - total_sale_last_month['total'])/total_sale_current_month['total'] * 100
    
    if sale_performance > 0:
        message = 'increased compare to the same period'
    if sale_performance < 0:
        message = 'decreased compare to the same period'
    
    top_ten_items = get_top_ten_data(timezone.now().replace(day=1,month=1), timezone.now())

    context = {
        'total_sale_month' : total_sale_month,
        'total_sale_pending' : total_sale_pending,
        'pending_paid_3days_ago' :  pending_paid_3days_ago,
        'achieve' : round(achieve),
        'short' : short_amount,
        'sale_trend' : sale_trend[::-1],
        'sale_day_current_month' : sale_day_current_month,
        'labels' : labels,
        'total_sale_current_month' : total_sale_current_month,
        'sale_day_last_month' : sale_day_last_month,
        'total_sale_last_month' : total_sale_last_month,
        'sale_performance' : round(sale_performance),
        'message' : message,
        'top_ten_items' : top_ten_items,
        'pending_count' : pending_count
    }

    return render(request, 'sale_transaction_statistic.html', context)

def sale_transaction_list_statistics_sale_performace(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        option = data.get('option')
        month_list = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        if option == 'week':
            end_date_current_month = timezone.now()
            start_date_current_month = end_date_current_month - timedelta(days=end_date_current_month.day-1) 
            sale_data_week_current_month, total_current_week = get_weekly_data(start_date_current_month,end_date_current_month,Sale)
            sale_week_current_month = []
            sale_week_current_month_label = []
            for item in sale_data_week_current_month:
                sale_week_current_month.append(item['total'])
                sale_week_current_month_label.append(item['week'])
            
            end_date_last_month = timezone.now() - relativedelta(months=1)
            start_date_last_month = end_date_last_month - timedelta(days=end_date_last_month.day-1)
            sale_data_week_last_month, total_last_week = get_weekly_data(start_date_last_month,end_date_last_month,Sale)
            sale_week_last_month = []
            sale_week_last_month_label = []
            for item in sale_data_week_last_month:
                sale_week_last_month.append(item['total'])
                sale_week_last_month_label.append(item['week'])
            
            if total_current_week['total'] == 0:
                sale_performance = 0
                message = ''
            else:
                sale_performance = (total_current_week['total'] - total_last_week['total'])/total_current_week['total'] * 100
            
            if sale_performance > 0:
                message = 'increased compare to the same period'
            if sale_performance < 0:
                message = 'decreased compare to the same period'

            items = {
                'data_current' : sale_week_current_month,
                'data_current_lebal' : month_list[end_date_current_month.month-1],
                'data_last' : sale_week_last_month,
                'data_last_label' : month_list[end_date_last_month.month-1],
                'labels' : sale_week_current_month_label,
                'sale_performance' : round(sale_performance),
                'message' : message
            }
            return JsonResponse(items, safe=False)
        if option == 'year':
            end_date_current_year = timezone.now()
            start_date_current_year = end_date_current_year.replace(day=1, month=1)
            sale_data_current_year, total_current_year = get_monthly_data(start_date_current_year,end_date_current_year,Sale)
            sale_month_current_year = []
            sale_month_current_year_label = []
            for item in sale_data_current_year:
                sale_month_current_year.append(item['total'])
                sale_month_current_year_label.append(item['month'])

            end_date_last_year = timezone.now()- relativedelta(years=1)
            start_date_last_year = end_date_last_year.replace(day=1, month=1)
            sale_data_last_year, total_last_year = get_monthly_data(start_date_last_year,end_date_last_year,Sale)
            sale_month_last_year = []
            sale_month_last_year_label = []
            for item in sale_data_last_year:
                sale_month_last_year.append(item['total'])
                sale_month_last_year_label.append(item['month'])

            if total_current_year['total'] == 0:
                sale_performance = 0
                message = ''
            else:
                sale_performance = (total_current_year['total'] - total_last_year['total'])/total_current_year['total'] * 100
            
            if sale_performance > 0:
                message = 'increased compare to the same period'
            if sale_performance < 0:
                message = 'decreased compare to the same period'

            items = {
                'data_current' : sale_month_current_year,
                'data_current_lebal' : end_date_current_year.year,
                'data_last' : sale_month_last_year,
                'data_last_label' : end_date_last_year.year,
                'labels' : sale_month_current_year_label,
                'sale_performance' : round(sale_performance),
                'message' : message
            }
            return JsonResponse(items, safe=False)

@receiver(post_save, sender=Purcase)
def purchase_transaction_save_item(sender, instance, created, **kwargs):
    if created:
        request = getattr(instance, '_request', None)
        try:
            with transaction.atomic():
                cart_item = json.loads(instance.cart_item)
                user = User.objects.get(username = request.user)
                for item in cart_item:
                    item['purchase_id'] = instance.id
                    item['unit_price'] = float(item['unit_price'])
                    del item['photo']
                    del item['name']
                    del item['cost']
                for data in cart_item:
                    sale_item = PurchaseItem(**data)
                    InventoryAdjustment.objects.create(name=Inventory.objects.get(id=data['item_id']), adjustment_type = 'In', type='Purchase', quantity = data['quantity'], store= Store.objects.get(name=user.userprofile.store), description = f'add purchase item {data['item_id']} purchase id {instance.id}')
                    sale_item.save()
                    messages.success(request, f'Purchase invoice has been recorded successfully.')
        except Exception as e:
            delete_sale = Purcase.objects.get(id=instance.id)
            delete_sale.delete()
            messages.error(request, f'Purchase invoice recorded fail. {e} ')

@in_groups('manager','superadmin')
def purchase_transaction(request):
    category_item = Category.objects.all()
    items = Inventory.objects.all().filter(status='Active')
    user = User.objects.get(username = request.user)
    promotion_code_def = PromotionCode.objects.get(name='no code')
    if request.method == 'POST':
        form = PurchaseForm(request.POST)
        cart_item = request.POST.get('cart_save_item')
        if form.is_valid():
            try:
                item = form.save(commit=False)
                item.cart_item = cart_item
                item._request = request
                item.user = request.user
                item.save()
                return redirect('purchase_transaction')   
            except Exception as e:
                messages.error(request,f'Purchase invoice recorded fail. {e}')
            except IntegrityError as e:
                messages.error(request,f'Purchase invoice recorded fail. {e}')      
    else:
        form = PurchaseForm(initial={'promotion_code': promotion_code_def})
        form1 = SupplierForm()
        form.fields['store'].initial = user.userprofile.store
        form1.fields['store'].initial = user.userprofile.store
    
    form = PurchaseForm(initial={'promotion_code': promotion_code_def})
    form1 = SupplierForm()
    form.fields['store'].initial = user.userprofile.store
    form1.fields['store'].initial = user.userprofile.store
    print(user.userprofile.store)
    context = {
        'category_item' : category_item,
        'items' : items,
        'form' : form,
        'form1' : form1
    }

    return render(request,'purchase_transaction.html',context)

@in_groups('manager','superadmin')
def purchase_transaction_add_supplier(request):
    if request.method == 'POST':
        form = SupplierForm(request.POST)
        if form.is_valid():
            try:
                form.save()
                last_customer = Supplier.objects.last()
                customer = list(Supplier.objects.values('first_name','last_name','id').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(id=last_customer.id))
                return JsonResponse({"message_sucess": "Supplier was added sucessfully.", 'customer' : customer })
            except Exception as e:
                return JsonResponse({"message_error": str(e)}) 
            except IntegrityError as e:
                return JsonResponse({"message_error": str(e)})   
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '            
        return JsonResponse({"form_error": error })       
    return JsonResponse({"error": "Invalid request"})

@in_groups('manager','superadmin')
def purchase_transaction_search_supplier(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        customer_list = list(Supplier.objects.values('first_name','last_name','id','phone_number1','phone_number2').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(Q(full_name__icontains=search_text) | Q(phone_number1__icontains=search_text) | Q(phone_number2__icontains=search_text)))
    return JsonResponse(customer_list, safe=False)

@in_groups('manager','superadmin')
def purchase_transaction_print_receipt(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        if (id):
            sale_id = list(Purcase.objects.values('id','store__name','store__address','store__logo', 'invoice_number','transaction_date','customer__first_name','customer__last_name','customer__address','customer__phone_number1','total_amount','discount','promotion_code__amount_discount','net_amount','payment_method','instruction').filter(id=id))
            sale_item = list(PurchaseItem.objects.values('item__name','quantity','unit_price','total_price').filter(purchase=id))
        else:
            last_record = Purcase.objects.last()
            sale_id = list(Purcase.objects.values('id','store__name','store__address','store__logo','invoice_number','transaction_date','customer__first_name','customer__last_name','customer__address','customer__phone_number1','total_amount','discount','promotion_code__amount_discount','net_amount','payment_method','instruction').filter(id=last_record.id))
            sale_item = list(PurchaseItem.objects.values('item__name','quantity','unit_price','total_price').filter(purchase=last_record.id))
    return JsonResponse(data={'sale_id' : sale_id, 'sale_item' : sale_item}, safe=False) 

@in_groups('manager','superadmin')
def purchase_transaction_list(request):
    now = timezone.now()
    category_item = Category.objects.all()
    items = Inventory.objects.all().filter(status='Active')
    purchase_transactions = Purcase.objects.all().filter(transaction_date__month = now.month).filter(deleted_at__isnull=True).order_by('-invoice_number')
    form = PurchaseForm()
    context = {
        'purchase_transactions' : purchase_transactions,
        'form' : form,
        'category_item' : category_item,
        'items' : items
    }
    return render(request, 'purchase_transaction_list.html', context)

@in_groups('manager','superadmin')
def purchase_transaction_list_filter_date(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        sale_transactions = list(Purcase.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)))
    return JsonResponse(sale_transactions, safe=False)

@in_groups('manager','superadmin')
def purchase_transaction_list_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        
        items = list(Purcase.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status', 'id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def purchase_transaction_list_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Purcase.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(Q(invoice_number__icontains=search_text) | Q(full_name__icontains=search_text)).order_by(f'{sort_method}{sortby}'))
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def purchase_transaction_list_delete_submit(request, id):
    if request.method == 'POST':
        sale_transaction = Purcase.objects.get(invoice_number=id)
        file_path = Path(settings.BASE_DIR / 'core/static/data')/'purchase_pending_delete.json'
        if file_path.exists():
            with open(file_path, 'r') as file:
                data = json.load(file)
        else: 
            data = {}
        
        data[sale_transaction.id] = ({
            'purchase_id' : sale_transaction.id,
            'status' : sale_transaction.status,
            'payment_method' : sale_transaction.payment_method
        })
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=4)
        
        if sale_transaction.status == 'Pending Delete':
            message = f'Invoice No. {id} was already sumitted for approval.'
            return JsonResponse({'message_error': message}, safe=False)
        else:
            sale_transaction.payment_method = 'Cash'    
            sale_transaction.status = 'Pending Delete'
            sale_transaction.save()
            message = f'Invoice No. {id} was successfully sumitted for approval.'
            return JsonResponse({'message_sucess': message}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def purchase_transaction_list_filter_payment(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        payment_method = data.get('payment_method')
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        items = list(Purcase.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status', 'id').filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(payment_method = payment_method))
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def purchase_transaction_list_approval(request):
    sale_transactions = Purcase.objects.all().filter(deleted_at__isnull=True).filter(status="Pending Delete").order_by('-invoice_number')
    context = {
        'sale_transactions' : sale_transactions
    }

    return render(request, 'purchase_transaction_list_approval.html', context)

@in_groups('manager','superadmin')
def purchase_transaction_list_filter_date_approval(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        sale_transactions = list(Purcase.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(status="Pending Delete"))
    return JsonResponse(sale_transactions, safe=False)    

@in_groups('manager','superadmin')
def purchase_transaction_list_sort_approval(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        
        items = list(Purcase.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status', 'id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(status="Pending Delete").order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def purchase_transaction_list_search_approval(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        date_from = date.fromisoformat(data.get('date_from'))
        date_to = date.fromisoformat(data.get('date_to'))
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Purcase.objects.values('store__name','invoice_number','transaction_date','user','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(status="Pending Delete").filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(Q(invoice_number__icontains=search_text) | Q(full_name__icontains=search_text)).order_by(f'{sort_method}{sortby}'))
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def purchase_transaction_list_delete_approval(request, id):
    if request.method == 'POST':
        with transaction.atomic():
            try:
                sale_transaction = Purcase.objects.get(invoice_number=id)
                sale_transaction.delete_custom()
                message = f'Invoice No. {id} was successfully deleted.'
                return JsonResponse({'message_sucess': message}, safe=False)
            except Exception as e:
                transaction.set_rollback(True) 
                return JsonResponse({'message_error': str(e)}, safe=False)
            except Exception as e:
                transaction.set_rollback(True) 
                return JsonResponse({'message_error': str(e)}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def purchase_transaction_list_delete_reject(request, id):
    if request.method == 'POST':
        sale_transaction = Purcase.objects.get(invoice_number=id)
        file_path = Path(settings.BASE_DIR /'core/static/data')/'purchase_pending_delete.json'
        with open(file_path, 'r') as file:
            data = json.load(file)
        if data[sale_transaction.id]['purchase_id'] == sale_transaction.id:
            sale_transaction.status = data[sale_transaction.id]['status']
            sale_transaction.payment_method = data[sale_transaction.id]['payment_method']            
            del data[sale_transaction.id]
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=4)
        sale_transaction.save()        
        message = f'Invoice No. {id} was successfully rejected.'
        return JsonResponse({'message_sucess': message}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def get_purchase_transaction(request, id):
    if request.method == 'POST':
        sale_transaction = list(Purcase.objects.values('store__name','invoice_number','transaction_date','user','customer__first_name','customer__last_name','total_amount','discount','promotion_code','net_amount','payment_method','status', 'id','promotion_code__amount_discount').filter(id=id))
        sale_item = list(PurchaseItem.objects.values('item__photo','item__name','quantity','unit_price','item__id').filter(purchase=id))
        for item in sale_item:
            if item['item__photo']:
                item['item__photo_url'] = f"{settings.MEDIA_URL}{item['item__photo']}"
            else:
                 item['item__photo_url'] = None   
        items = {
            'sale_transaction' : sale_transaction,
            'sale_item' : sale_item,
        }
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def purchase_transaction_save(request):
    if request.method == "POST":
        data = json.loads(request.body)
        cart = data.get('cart')
        invoice_number_id = data.get('invoice_number_id')
        invoice_number = data.get('invoice_number')
        user = User.objects.get(username = request.user)
        sale_ids = [item['item_id'] for item in cart]
        with transaction.atomic():    
            try:
                remove_sale = PurchaseItem.objects.filter(purchase=invoice_number_id).exclude(item__in=sale_ids)
                for item in remove_sale:
                    InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item.item_id), adjustment_type = 'Out', type='Adjustment', quantity = item.quantity, store=Store.objects.get(name=user.userprofile.store), description = f'remove purchase item {item.item_id} sale id {item.purchase}')
                    item.delete()
            except Exception as e:
                transaction.set_rollback(True)
                return JsonResponse({'message_error': str(e)}, safe=False)
            for item in cart:
                if PurchaseItem.objects.filter(purchase=invoice_number_id).filter(item=item['item_id']).exists():
                    try:
                        update_sale_item = PurchaseItem.objects.filter(purchase=invoice_number_id).get(item=item['item_id'])
                        different_item_quantiry = item['quantity'] - update_sale_item.quantity
                        if different_item_quantiry > 0:
                            InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'In', type='Purchase', quantity = different_item_quantiry, store=Store.objects.get(name=user.userprofile.store), description = f'increase quantity purchase item {item['item_id']} purchase id {invoice_number_id}')
                        if different_item_quantiry < 0:
                            InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'Out', type='Adjustment', quantity = abs(different_item_quantiry), store=Store.objects.get(name=user.userprofile.store), description = f'decrease quantity purchase item {item['item_id']} purchase id {invoice_number_id}')    
                        update_sale_item.quantity = item['quantity']
                        update_sale_item.unit_price = float(item['unit_price'])
                        update_sale_item.save()    
                    except Exception as e:
                        transaction.set_rollback(True) 
                        return JsonResponse({'message_error': str(e)}, safe=False)
                else:   
                    try:
                        sale=Purcase.objects.get(id=invoice_number_id)
                        new_sale_item = Inventory.objects.get(id=item['item_id'])
                        InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'In', type='Purchase', quantity = item['quantity'], store=Store.objects.get(name=user.userprofile.store), description = f'add purchase item {item['item_id']} purchase id {invoice_number_id}')        
                        PurchaseItem.objects.create(purchase=sale, item=new_sale_item, quantity= item['quantity'], unit_price=(item['unit_price']))
                    except Exception as e:
                        transaction.set_rollback(True) 
                        return JsonResponse({'message_error': str(e)}, safe=False)        
            try:    
                Purcase.objects.get(id=invoice_number_id).save()
                message = f'Invoice No. {invoice_number} was successfully saved.'
                return JsonResponse({'message_sucess': message}, safe=False) 
            except Exception as e:
                return JsonResponse({'message_error': str(e)}, safe=False)

@in_groups('manager','superadmin')         
def purchase_transaction_paid(request):
    if request.method == "POST":
        data = json.loads(request.body)
        cart = data.get('cart')
        invoice_number_id = data.get('invoice_number_id')
        invoice_number = data.get('invoice_number')
        user = User.objects.get(username = request.user)
        sale_ids = [item['item_id'] for item in cart]
        with transaction.atomic():    
            try:
                remove_sale = PurchaseItem.objects.filter(purchase=invoice_number_id).exclude(item__in=sale_ids)
                for item in remove_sale:
                    InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item.item_id), adjustment_type = 'Out', type='Adjustment', quantity = item.quantity, store=Store.objects.get(name=user.userprofile.store), description = f'remove purchase item {item.item_id} sale id {item.purchase}')
                    item.delete()
            except Exception as e:
                transaction.set_rollback(True)
                return JsonResponse({'message_error': str(e)}, safe=False)
            for item in cart:
                if PurchaseItem.objects.filter(purchase=invoice_number_id).filter(item=item['item_id']).exists():
                    try:
                        update_sale_item = PurchaseItem.objects.filter(purchase=invoice_number_id).get(item=item['item_id'])
                        different_item_quantiry = item['quantity'] - update_sale_item.quantity
                        if different_item_quantiry > 0:
                            InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'In', type='Purchase', quantity = different_item_quantiry, store=Store.objects.get(name=user.userprofile.store), description = f'increase quantity purchase item {item['item_id']} purchase id {invoice_number_id}')
                        if different_item_quantiry < 0:
                            InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'Out', type='Adjustment', quantity = abs(different_item_quantiry), store=Store.objects.get(name=user.userprofile.store), description = f'decrease quantity purchase item {item['item_id']} purchase id {invoice_number_id}')    
                        update_sale_item.quantity = item['quantity']
                        update_sale_item.unit_price = float(item['unit_price'])
                        update_sale_item.save()    
                    except Exception as e:
                        transaction.set_rollback(True) 
                        return JsonResponse({'message_error': str(e)}, safe=False)
                else:   
                    try:
                        sale=Purcase.objects.get(id=invoice_number_id)
                        new_sale_item = Inventory.objects.get(id=item['item_id'])
                        InventoryAdjustment.objects.create(name=Inventory.objects.get(id=item['item_id']), adjustment_type = 'In', type='Purchase', quantity = item['quantity'], store=Store.objects.get(name=user.userprofile.store), description = f'add purchase item {item['item_id']} purchase id {invoice_number_id}')        
                        PurchaseItem.objects.create(purchase=sale, item=new_sale_item, quantity= item['quantity'], unit_price=float(item['unit_price']))
                    except Exception as e:
                        transaction.set_rollback(True) 
                        return JsonResponse({'message_error': str(e)}, safe=False)         
        sale_transaction = Purcase.objects.get(id=invoice_number_id)
        sale_transaction.promotion_code = PromotionCode.objects.get(id=data.get('promotion_code'))
        sale_transaction.status = 'Paid'
        sale_transaction.discount = data.get('discount')
        sale_transaction.payment_method = data.get('payment_method')
        sale_transaction.save()
        message = f'Invoice No. {invoice_number} was successfully paid.'
        return JsonResponse({'message_sucess': message}, safe=False) 
    
def dashborad(request):
    total_sale_month = Sale.objects.filter(status='Paid').filter(deleted_at__isnull=True).filter(transaction_date__date__range=(timezone.now().replace(day=1),timezone.now())).aggregate(total=Sum('total_amount'))['total'] or 0
    total_sale_pending = Sale.objects.filter(status='Pending').filter(deleted_at__isnull=True).aggregate(total=Sum('total_amount'))['total'] or 0
    pending_count = Sale.objects.filter(status='Pending').filter(deleted_at__isnull=True).count()
    pending_paid_3days_ago = Sale.objects.filter(status='Pending').filter(deleted_at__isnull=True).filter(transaction_date__date__lte=(timezone.now() - timedelta(days=3))).aggregate(total=Sum('total_amount'))['total'] or 0    
    sale_target = Store.objects.get(id=1)
    days_month = calendar.monthrange(timezone.now().year, timezone.now().month)
    pro_target = (sale_target.sale_target/days_month[1])*int(timezone.now().day) 
    achieve = (Decimal(total_sale_month) / (pro_target))*100
    short_amount = Decimal(sale_target.sale_target) - total_sale_month

    data_sale_day_current_month, total_sale_current_month = get_daily_data(timezone.now().replace(day=1), timezone.now(), Sale)
    sale_day_current_month = []
    labels = []
    for item in data_sale_day_current_month:
        sale_day_current_month.append(item['total'])
        labels.append(item['day'])

    data_sale_day_last_month, total_sale_last_month = get_daily_data((timezone.now()-relativedelta(months=1)).replace(day=1), timezone.now()- relativedelta(months=1),Sale)
    sale_day_last_month = []
    for item in data_sale_day_last_month:
        sale_day_last_month.append(item['total'])
   
    if total_sale_current_month['total'] == 0:
        sale_performance = 0
        message = ''
    else:
        sale_performance = (total_sale_current_month['total'] - total_sale_last_month['total'])/total_sale_current_month['total'] * 100
    
    if sale_performance > 0:
        message = 'increased compare to the same period'
    if sale_performance < 0:
        message = 'decreased compare to the same period'
   
    top_ten_items = get_top_ten_data(timezone.now().replace(day=1,month=1), timezone.now())
    top_ten_customer = get_top_ten_customer(timezone.now().replace(day=1,month=1), timezone.now())

    low_stock_item = Inventory.objects.filter(quantity__lt=F('reorder_alert')).filter(status='Active')

    data_sale_trend, total = get_trend_daily(7,Sale)
    sale_trend = []
    for item in data_sale_trend:
        sale_trend.append(item['total'])
    
    profit_data = get_profit(timezone.now().replace(day=1), timezone.now())
    if profit_data == 0:
        gross_profit_margin = 0
        profit = 0
    else:
        gross_profit_margin = round(((profit_data['sale'] - profit_data['cost'])/profit_data['sale']) * 100)
        profit = profit_data['profit']
    
    userprofile = UserProfile.objects.get(user__username = request.user)
    password_not_change_number_of_date = timezone.now().date() - userprofile.password_change_date
    passwor_warning = None
    if password_not_change_number_of_date.days >=30:
        passwor_warning = 'Your password has not been changed for more than 30 days. Please change your password as soon as possible.'

    context = {
       'total_sale_month' : total_sale_month,
       'total_sale_pending' : total_sale_pending,
       'pending_paid_3days_ago' : pending_paid_3days_ago, 
       'achieve' : round(achieve),
       'short' : short_amount,
       'sale_day_current_month' : sale_day_current_month,
       'labels' : labels,
       'total_sale_current_month' : total_sale_current_month,
       'sale_day_last_month' : sale_day_last_month,
       'total_sale_last_month' : total_sale_last_month,
       'sale_performance' : round(sale_performance),
       'message' : message,
       'top_ten_items' : top_ten_items,
       'top_ten_customer' : top_ten_customer,
       'low_stock_item' : low_stock_item,
       'sale_trend' : sale_trend[::-1],
       'profit' : profit,
       'gross_profit_margin' : gross_profit_margin,
       'pending_count' : pending_count,
       'passwor_warning' : passwor_warning
   }
    
    return render(request, 'dashborad.html', context)

@in_groups('manager','superadmin')
def supplier_management(request):
    items = Supplier.objects.all().filter(status='Active')
    user = User.objects.get(username = request.user)
    if request.method == 'POST':
        form = SupplierForm(request.POST, request.FILES)
        if form.is_valid():            
            try:
                form.save()
                messages.success(request,'Supplier was created sucessfully.')
                return redirect ('supplier_management')
            except Exception as e:
                messages.error(request,e)
                return redirect ('supplier_management')
            except IntegrityError as e:
                messages.error(request,e)
                return redirect ('supplier_management')
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '            
            messages.error(request,error)
            return redirect ('supplier_management') 
    else:
        form = SupplierForm()
        form.fields['store'].initial = user.userprofile.store
    
    context = {
        'items' : items,
        'form' : form,
    }
    return render(request, 'supplier_management.html', context)

@in_groups('manager','superadmin')
def supplier_management_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        items = list(Supplier.objects.values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','note','created_on').filter(status='Active').order_by(f'{sort_method}{sortby}'))

    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def supplier_management_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Supplier.objects.values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','note','created_on').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(Q(full_name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Active').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def get_supplier(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('product_id')
        request.session['temp_product_id'] = product_id 
        item = list(Supplier.objects.values('id','store__id','first_name','last_name','gender','phone_number1','e_mail','membership','note','created_on','address').filter(id=product_id))  
         
    return JsonResponse(item, safe=False)

@in_groups('manager','superadmin')
def save_supplier(request, product_id):
    if request.method == 'POST':
        update_item = Supplier.objects.get(id = product_id)
        form = SupplierForm(request.POST, request.FILES, instance=update_item)
        if form.is_valid():
            try:
                form.save()
                return JsonResponse({"message_sucess": "Supplier was updated sucessfully."})
            except Exception as e:
                return JsonResponse({"message_error": str(e)})
            except IntegrityError as e:
                return JsonResponse({"message_error": str(e)})
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '
        return JsonResponse({"message_error": str(error)})       
    return JsonResponse({"error": "Invalid request"})

@in_groups('manager','superadmin')
def deactivate_supplier(request, product_id):
    if request.method == 'POST':
        item = Supplier.objects.get(id=product_id)
        try:
            customer_balance = item.credit_balance.all().order_by('created_at').last().balance
        except:
            customer_balance = 0
        if customer_balance == 0:
            item.status = 'Inactive'
            item.save()
            message = f'Supplier with ID No. {product_id} was successfully deactivated.'
            return JsonResponse({'message_sucess': message}, safe=False)
        else:
            message = f'customer with Id No. {product_id} remains {customer_balance} credit balance. Please make sure the credit balance of customer is zeroized before perform the action.'
            return JsonResponse({'message_error': message}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def deactivate_supplier_list(request):
    items = Supplier.objects.all().filter(status='Inactive')
    if request.method == 'POST':
        form = CategoryForm(request.POST, request.FILES)
        if form.is_valid():            
            form.save()
            messages.success(request,'supplier was created sucessfully.')
            return redirect ('supplier_management')
    else:
        form = SupplierForm()
    
    context = {
        'items' : items,
        'form' : form,
    }
    
    return render(request, 'supplier_management_deactivate.html', context)

@in_groups('manager','superadmin')
def supplier_management_deactivate_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        items = list(Supplier.objects.values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','note','created_on').filter(status='Inactive').order_by(f'{sort_method}{sortby}'))

    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def supplier_management_deactivate_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Supplier.objects.values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','note','created_on').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(Q(full_name__icontains=search_text) | Q(id__icontains=search_text)).filter(status='Inactive').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def activate_supplier(request, product_id):
    if request.method == 'POST':
        try:
            item = Supplier.objects.get(id=product_id)      
            item.status = 'Active'
            item.save()
            message = f'Supplier with ID No. {product_id} was successfully activated.'
            return JsonResponse({'message_sucess': message}, safe=False)
        except Exception as e:
            return JsonResponse({'message_error': str(e)}, safe=False)
        except IntegrityError as e:
            return JsonResponse({'message_error': str(e)}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def expanse_management(request):
    items = ExpanseItem.objects.all().filter(deleted_at__isnull=True)
    user = User.objects.get(username = request.user)
    if request.method == 'POST':
        form = ExpanseForm(request.POST, request.FILES)
        if form.is_valid():            
            try:
                form.save()
                messages.success(request,'Expanse was created sucessfully.')
                return redirect ('expanse_management')
            except Exception as e:
                messages.error(request, str(e))
            except IntegrityError as e:
                messages.error(request, str(e))
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '            
        messages.error(request, error)
           
    else:
        form = ExpanseForm()
        form.fields['store'].initial = user.userprofile.store
            
    context = {
        'items' : items,
        'form' : form,
    }
    
    return render(request, 'expanse_management.html', context)

@in_groups('manager','superadmin')
def expanse_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        
        items = list(ExpanseItem.objects.values('id','description','store__name','category__name','amount','transaction_date').filter(deleted_at__isnull=True).order_by(f'{sort_method}{sortby}'))    
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def expanse_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(ExpanseItem.objects.values('id','description','store__name','category__name','amount','transaction_date').filter(deleted_at__isnull=True).filter(Q(description__icontains=search_text) | Q(id__icontains=search_text)).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def expanse_get(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('product_id')
        request.session['temp_product_id'] = product_id 
        item = list(ExpanseItem.objects.values('id','description','store','category','amount','transaction_date', 'reference_photo').filter(id=product_id))
    return JsonResponse(item, safe=False)

@in_groups('manager','superadmin')
def expanse_save(request, id):
    if request.method == 'POST':
        update_item = ExpanseItem.objects.get(id = id)
        form = ExpanseForm(request.POST, request.FILES, instance=update_item)
        if form.is_valid():
            try:
                form.save()
                return JsonResponse({"message_sucess": "Expanse was updated sucessfully."})
            except Exception as e:
                return JsonResponse({"message_error": str(e)})
            except Exception as e:
                return JsonResponse({"message_error": str(e)})
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '
            return JsonResponse({"message_error": error})       
    return JsonResponse({"error": "Invalid request"})

@in_groups('manager','superadmin')
def expanse_delete(request, id):
    if request.method == 'POST':
        item = ExpanseItem.objects.get(id = id)
        item.delete()
        return JsonResponse({"message": "Expanse was deleted sucessfully."})
    return JsonResponse({"error": "Invalid request"})

@in_groups('manager','superadmin')
def expanse_category(request):
    items_category = ExpanseCategory.objects.all()
    if request.method == 'POST':
        form = ExpanseCategoryForm(request.POST)
        if form.is_valid():            
            try:
                form.save()
                messages.success(request, 'Category was created sucessfully.')
                return redirect ('expanse_category')   
            except Exception as e:
                messages.error(request, str(e))
            except Exception as e:
                messages.error(request, str(e))
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '
        messages.error(request, error)
    
    else:
        form = ExpanseCategoryForm() 
    
    context = {
        'items' : items_category,
        'form' : form,
    }

    return render(request, 'expanse_category.html', context)

@in_groups('manager','superadmin')
def expanse_category_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        items = list(ExpanseCategory.objects.values('id','name').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def expanse_category_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(ExpanseCategory.objects.values('id','name').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def expanse_category_get(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        product_id = data.get('product_id')
        request.session['temp_product_id'] = product_id 
        item = list(ExpanseCategory.objects.values('id','name').filter(id=product_id))   
    return JsonResponse(item, safe=False)

@in_groups('manager','superadmin')
def expanse_category_modify(request, id):
    if request.method == 'POST':
        update_category = ExpanseCategory.objects.get(name = id)
        form = ExpanseCategoryForm(request.POST, instance=update_category)
        if form.is_valid():
            form.save()
            return JsonResponse({"message": "Category was updated sucessfully."})
        return JsonResponse({"error": form.errors})       
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def expanse_category_filter(request, name):
    user = User.objects.get(username = request.user)
    if ExpanseItem.objects.all().filter(deleted_at__isnull=True).filter(category__name=name).exists():
        items = ExpanseItem.objects.all().filter(deleted_at__isnull=True).filter(category__name=name)
        category=None     
    else:
        category = name
        items = None
    
    form = ExpanseForm()
    form.fields['store'].initial = user.userprofile.store
    
    context = {
        'items' : items,
        'form' : form,
        'category' : category,     
    }
    return render(request, 'expanse_category_filter.html', context)

@in_groups('manager','superadmin')
def expanse_category_filter_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        expanse_category = data.get('expanse_category')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        
        items = list(ExpanseItem.objects.values('id','description','store__name','category__name','amount','transaction_date').filter(deleted_at__isnull=True).filter(category__name=expanse_category).order_by(f'{sort_method}{sortby}'))    
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def expanse_category_filter_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        expanse_category = data.get('expanse_category')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(ExpanseItem.objects.values('id','description','store__name','category__name','amount','transaction_date').filter(deleted_at__isnull=True).filter(Q(description__icontains=search_text) | Q(id__icontains=search_text)).filter(category__name=expanse_category).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

def report_management(request):
    
    return render(request, 'report_management.html')

def report_management_sale_by_invoice(request):
   
    return render(request, 'report_management_sale_by_invoice.html')

def report_management_sale_by_invoice_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('start_date'))
        date_to = date.fromisoformat(data.get('end_date'))
        store = []
        for item in Store.objects.all().filter(is_selected = True):
            store.append(item.name)
        sale_transactions = (Sale.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','promotion_code__amount_discount','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)))
        if (sale_transactions):
            sub_total = sale_transactions.values('store','store__name').annotate(total_sale=Sum('total_amount'), total_discount= Sum('discount_amount'),total_net_sale = Sum('net_amount'), total_promotion_code = Sum('promotion_code__amount_discount')).order_by('store')
            items={
                'sale_transactions' : list(sale_transactions),
                'sub_total' : list(sub_total),
                'store' : store
                }
        else :
            items={
                'message' : 'Record not found!'
            }
    
    return JsonResponse(items, safe=False)  

def report_management_sale_by_invoice_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_to = date.fromisoformat(data.get('end_date'))
        date_from = date.fromisoformat(data.get('start_date'))    
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sale Report By Invoice'
        ws.merge_cells('A2:K2')
        ws.merge_cells('A3:K3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Sale Report By Invoice'
        ws['A3'] = f'From {date_from} to {date_to}'
        ws['A4'] = ''
        ws.append(['DATE','STORE','INVOICE No.','USER','CUSTOMER','TOTAL','PROMOTION CODE','DISCOUNT(%)','NET TOTAL','PAYMENT METHOD','STATUS'])    
        sale_transactions = (Sale.objects.values('transaction_date','store__name','invoice_number','user__username','customer','total_amount','promotion_code__name','discount','net_amount','payment_method','status').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).order_by('invoice_number'))
        for sale in sale_transactions:
            ws.append([sale['transaction_date'].strftime('%Y-%m-%d %H:%M:%S'), sale['store__name'],sale['invoice_number'],sale['user__username'],sale['full_name'].capitalize(), sale['total_amount'],sale['promotion_code__name'],sale['discount'],sale['net_amount'], sale['payment_method'], sale['status']])          
       
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

def report_management_sale_by_date(request):
    
    return render(request, 'report_management_sale_by_date.html')

def report_management_sale_by_date_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('start_date'))
        date_to = date.fromisoformat(data.get('end_date'))
        store = []
        for item in Store.objects.all().filter(is_selected = True):
            store.append(item.name)  
        sale_transactions = (Sale.objects.filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).annotate(sale_date= TruncDate('transaction_date'), discount_amount= (F('total_amount') * F('discount')/100)).values('store__name','sale_date','promotion_code__amount_discount').annotate(total_sale=Sum('total_amount'), total_net_sale=Sum('net_amount'),total_discount=Sum('discount_amount'), total_promotion_code=Sum('promotion_code__amount_discount')))
        if (sale_transactions):
            sub_total = sale_transactions.values('store','store__name').annotate(total_sale=Sum('total_amount'), total_discount= Sum('discount_amount'),total_net_sale = Sum('net_amount'), total_promotion_code = Sum('promotion_code__amount_discount')).order_by('store')
            items={
                'sale_transactions' : list(sale_transactions), 
                'sub_total' : list(sub_total),
                'store' : store      
            }
        else:
            items={
                'message' : 'Record not found!'
            }
    return JsonResponse(items, safe=False)

def report_management_sale_by_date_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_to = date.fromisoformat(data.get('end_date'))
        date_from = date.fromisoformat(data.get('start_date'))    
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sale Report By Date'
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Sale Report By Date'
        ws['A3'] = f'From {date_from} to {date_to}'
        ws['A4'] = ''
        ws.append(['DATE','STORE','TOTAL','PROMOTION CODE','DISCOUNT','NET TOTAL'])    
        sale_transactions = (Sale.objects.filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).annotate(sale_date= TruncDate('transaction_date'), discount_amount= (F('total_amount') * F('discount')/100)).values('store__name','sale_date').annotate(total_sale=Sum('total_amount'), total_net_sale=Sum('net_amount'),total_discount=Sum('discount_amount'), total_promotion_code=Sum('promotion_code__amount_discount')))
        for sale in sale_transactions:
            ws.append([sale['sale_date'].strftime('%Y-%m-%d'), sale['store__name'],sale['total_sale'],sale['total_promotion_code'],sale['total_discount'],sale['total_net_sale']])         
       
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

def report_management_sale_by_user(request):
    form = SaleForm()
    form.fields['user'].queryset = UserProfile.objects.all()

    context = {
        'form' : form
    }

    return render(request, 'report_management_sale_by_user.html', context)

def report_management_sale_by_user_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('start_date'))
        date_to = date.fromisoformat(data.get('end_date'))
        user = UserProfile.objects.get(pk=data.get('user'))
        try:    
            sale_transactions = Sale.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(user=user.user).order_by('invoice_number')
            total_sale= sale_transactions.aggregate(total_sale=Sum('total_amount'))
            total_discount = sale_transactions.aggregate(total_discount=Sum('discount_amount'))
            total_net_sale = sale_transactions.aggregate(total_net_sale=Sum('net_amount'))
            items={
                'sale_transactions' : list(sale_transactions),
                'total_sale' : Decimal(total_sale['total_sale']).quantize(Decimal('0.00')),
                'total_discount' : Decimal(total_discount['total_discount']).quantize(Decimal('0.00')),
                'total_net_sale' : Decimal(total_net_sale['total_net_sale']).quantize(Decimal('0.00'))
            }
        except:
            items={
                'message' : 'Record not found!'
            }
        return JsonResponse(items, safe=False)

def report_management_sale_by_user_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_to = date.fromisoformat(data.get('end_date'))
        date_from = date.fromisoformat(data.get('start_date'))   
        user = UserProfile.objects.get(pk=data.get('user')) 
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sale Report By User'
        ws.merge_cells('A2:K2')
        ws.merge_cells('A3:K3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Sale Report By User'
        ws['A3'] = f'From {date_from} to {date_to}'
        ws['A4'] = ''
        ws.append(['DATE','STORE','INVOICE No.','USER','CUSTOMER','TOTAL','PROMOTION CODE','DISCOUNT(%)','NET TOTAL','PAYMENT METHOD','STATUS'])    
        sale_transactions = Sale.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(user=user.user).order_by('invoice_number')
        for sale in sale_transactions:
            ws.append([sale['transaction_date'].strftime('%Y-%m-%d %H:%M:%S'), sale['store__name'],sale['invoice_number'],sale['user__username'],sale['full_name'].capitalize(), sale['total_amount'],sale['promotion_code__name'],sale['discount'],sale['net_amount'], sale['payment_method'], sale['status']])          
       
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

def report_management_sale_by_payment_method(request):
    form = SaleForm()
    form['payment_method'].initial = 'Cash'

    context = {
        'form' : form
    }

    return render(request, 'report_management_sale_by_payment_method.html', context)

def report_management_sale_by_payment_method_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('start_date'))
        date_to = date.fromisoformat(data.get('end_date'))
        payment_method = data.get('payment_method')
        store = []
        for item in Store.objects.all().filter(is_selected = True):
            store.append(item.name)
        sale_transactions = Sale.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(payment_method=payment_method).order_by('invoice_number')
        if (sale_transactions):
            sub_total = sale_transactions.values('store','store__name').annotate(total_sale=Sum('total_amount'), total_discount= Sum('discount_amount'),total_net_sale = Sum('net_amount'), total_promotion_code = Sum('promotion_code__amount_discount')).order_by('store')
            items={
                'sale_transactions' : list(sale_transactions),
                'sub_total' : list(sub_total),
                'store' : store
            }
        else:
            items={
                'message' : 'Record not found!'
            }
        return JsonResponse(items, safe=False)

def report_management_sale_by_payment_method_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_to = date.fromisoformat(data.get('end_date'))
        date_from = date.fromisoformat(data.get('start_date'))   
        payment_method = data.get('payment_method') 
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sale Report By Payment Method'
        ws.merge_cells('A2:K2')
        ws.merge_cells('A3:K3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Sale Report By Payment Method'
        ws['A3'] = f'From {date_from} to {date_to}'
        ws['A4'] = ''
        ws.append(['DATE','STORE','INVOICE No.','USER','CUSTOMER','TOTAL','PROMOTION CODE','DISCOUNT(%)','NET TOTAL','PAYMENT METHOD','STATUS'])    
        sale_transactions = Sale.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(payment_method=payment_method).order_by('invoice_number')
        for sale in sale_transactions:
            ws.append([sale['transaction_date'].strftime('%Y-%m-%d %H:%M:%S'), sale['store__name'],sale['invoice_number'],sale['user__username'],sale['full_name'].capitalize(), sale['total_amount'],sale['promotion_code__name'],sale['discount'],sale['net_amount'], sale['payment_method'], sale['status']])          
       
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response
    
def report_management_sale_by_status(request):
    form = SaleForm()
    form['status'].initial = 'Paid'

    context = {
        'form' : form
    }

    return render(request, 'report_management_sale_by_status.html', context)

def report_management_sale_by_status_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('start_date'))
        date_to = date.fromisoformat(data.get('end_date'))
        status = data.get('status')
        store = []
        for item in Store.objects.all().filter(is_selected = True):
            store.append(item.name)    
        if status == "Deleted":
            sale_transactions = Sale.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(transaction_date__date__range = (date_from,date_to)).filter(status=status).order_by('invoice_number')
        else:
            sale_transactions = Sale.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(status=status).order_by('invoice_number')
            
        if sale_transactions:
            sub_total = sale_transactions.values('store','store__name').annotate(total_sale=Sum('total_amount'), total_discount= Sum('discount_amount'),total_net_sale = Sum('net_amount'), total_promotion_code = Sum('promotion_code__amount_discount')).order_by('store')
            items={
                'sale_transactions' : list(sale_transactions),
                'sub_total' : list(sub_total),
                'store' : store
            }
        else:
            items={
                'message' : 'Record not found!'
            }

    return JsonResponse(items, safe=False) 

def report_management_sale_by_status_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_to = date.fromisoformat(data.get('end_date'))
        date_from = date.fromisoformat(data.get('start_date'))   
        status = data.get('status') 
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sale Report By Status'
        ws.merge_cells('A2:K2')
        ws.merge_cells('A3:K3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Sale Report By Status'
        ws['A3'] = f'From {date_from} to {date_to}'
        ws['A4'] = ''
        ws.append(['DATE','STORE','INVOICE No.','USER','CUSTOMER','TOTAL','PROMOTION CODE','DISCOUNT(%)','NET TOTAL','PAYMENT METHOD','STATUS'])    
        if status == "Deleted":
            sale_transactions = Sale.objects.values('store__name','invoice_number','transaction_date','user_username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(transaction_date__date__range = (date_from,date_to)).filter(status=status).order_by('invoice_number')
        else:
            sale_transactions = Sale.objects.values('store__name','invoice_number','transaction_date','user_username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(status=status).order_by('invoice_number')
        for sale in sale_transactions:
            ws.append([sale['transaction_date'].strftime('%Y-%m-%d %H:%M:%S'), sale['store__name'],sale['invoice_number'],sale['user_username'],sale['full_name'].capitalize(), sale['total_amount'],sale['promotion_code__name'],sale['discount'],sale['net_amount'], sale['payment_method'], sale['status']])          
       
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')
def report_management_purchase_by_invoice(request):
    
    return render(request, 'report_management_purchase_by_invoice.html')

@in_groups('manager','superadmin')
def report_management_purchase_by_invoice_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('start_date'))
        date_to = date.fromisoformat(data.get('end_date'))
        store = []
        for item in Store.objects.all().filter(is_selected = True):
            store.append(item.name)  
        sale_transactions = (Purcase.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).order_by('invoice_number'))
        if sale_transactions:
            sub_total = sale_transactions.values('store','store__name').annotate(total_sale=Sum('total_amount'), total_discount= Sum('discount_amount'),total_net_sale = Sum('net_amount'), total_promotion_code = Sum('promotion_code__amount_discount')).order_by('store')
            items={
                'sale_transactions' : list(sale_transactions),
                'sub_total' : list(sub_total),
                'store' : store
            }
        else:
            items={
                'message' : 'Record not found!'
            }
    return JsonResponse(items, safe=False)  

@in_groups('manager','superadmin')
def report_management_purchase_by_invoice_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_to = date.fromisoformat(data.get('end_date'))
        date_from = date.fromisoformat(data.get('start_date'))    
        wb = Workbook()
        ws = wb.active
        ws.title = 'Purchase Report By Invoice'
        ws.merge_cells('A2:K2')
        ws.merge_cells('A3:K3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Sale Report By Invoice'
        ws['A3'] = f'From {date_from} to {date_to}'
        ws['A4'] = ''
        ws.append(['DATE','STORE','INVOICE No.','USER','CUSTOMER','TOTAL','PROMOTION CODE','DISCOUNT(%)','NET TOTAL','PAYMENT METHOD','STATUS'])    
        sale_transactions = (Purcase.objects.values('transaction_date','store__name','invoice_number','user__username','customer','total_amount','promotion_code__name','discount','net_amount','payment_method','status').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name')).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).order_by('invoice_number'))
        for sale in sale_transactions:
            ws.append([sale['transaction_date'].strftime('%Y-%m-%d %H:%M:%S'), sale['store__name'],sale['invoice_number'],sale['user__username'],sale['full_name'].capitalize(), sale['total_amount'],sale['promotion_code__name'],sale['discount'],sale['net_amount'], sale['payment_method'], sale['status']])          
       
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')    
def report_management_purchase_by_date(request):
    
    return render(request, 'report_management_purchase_by_date.html')

@in_groups('manager','superadmin')
def report_management_purchase_by_date_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('start_date'))
        date_to = date.fromisoformat(data.get('end_date'))
        store = []
        for item in Store.objects.all().filter(is_selected = True):
            store.append(item.name)    
        sale_transactions = (Purcase.objects.filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).annotate(sale_date= TruncDate('transaction_date'), discount_amount= (F('total_amount') * F('discount')/100)).values('store__name','sale_date').annotate(total_sale=Sum('total_amount'), total_net_sale=Sum('net_amount'),total_discount=Sum('discount_amount'), total_promotion_code=Sum('promotion_code__amount_discount')))
        if sale_transactions:
            sub_total = sale_transactions.values('store','store__name').annotate(total_sale=Sum('total_amount'), total_discount= Sum('discount_amount'),total_net_sale = Sum('net_amount'), total_promotion_code = Sum('promotion_code__amount_discount')).order_by('store')
            items={
                'sale_transactions' : list(sale_transactions),
                'sub_total' : list(sub_total),
                'store' : store
            }
        else:
            items={
                'message' : 'Record not found!'
            }
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def report_management_purchase_by_date_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_to = date.fromisoformat(data.get('end_date'))
        date_from = date.fromisoformat(data.get('start_date'))    
        wb = Workbook()
        ws = wb.active
        ws.title = 'Purchase Report By Date'
        ws.merge_cells('A2:F2')
        ws.merge_cells('A3:F3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Purchase Report By Date'
        ws['A3'] = f'From {date_from} to {date_to}'
        ws['A4'] = ''
        ws.append(['DATE','STORE','TOTAL','PROMOTION CODE','DISCOUNT','NET TOTAL'])    
        sale_transactions = (Purcase.objects.filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).annotate(sale_date= TruncDate('transaction_date'), discount_amount= (F('total_amount') * F('discount')/100)).values('store__name','sale_date').annotate(total_sale=Sum('total_amount'), total_net_sale=Sum('net_amount'),total_discount=Sum('discount_amount'), total_promotion_code=Sum('promotion_code__amount_discount')))
        for sale in sale_transactions:
            ws.append([sale['sale_date'].strftime('%Y-%m-%d'), sale['store__name'],sale['total_sale'],sale['total_promotion_code'],sale['total_discount'],sale['total_net_sale']])         
       
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')
def report_management_purchase_by_user(request):
    form = SaleForm()
    form.fields['user'].queryset = UserProfile.objects.all()

    context = {
        'form' : form
    }

    return render(request, 'report_management_purchase_by_user.html', context)

@in_groups('manager','superadmin')
def report_management_purchase_by_user_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('start_date'))
        date_to = date.fromisoformat(data.get('end_date'))
        user = UserProfile.objects.get(pk=data.get('user'))
        try:    
            sale_transactions = Purcase.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(user=user.user).order_by('invoice_number')
            total_sale= sale_transactions.aggregate(total_sale=Sum('total_amount'))
            total_discount = sale_transactions.aggregate(total_discount=Sum('discount_amount'))
            total_net_sale = sale_transactions.aggregate(total_net_sale=Sum('net_amount'))
            items={
                'sale_transactions' : list(sale_transactions),
                'total_sale' : Decimal(total_sale['total_sale']).quantize(Decimal('0.00')),
                'total_discount' : Decimal(total_discount['total_discount']).quantize(Decimal('0.00')),
                'total_net_sale' : Decimal(total_net_sale['total_net_sale']).quantize(Decimal('0.00'))
            }
        except Exception as e:
            items={
                'message' : 'Record not found!'
            }
        return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def report_management_purchase_by_user_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_to = date.fromisoformat(data.get('end_date'))
        date_from = date.fromisoformat(data.get('start_date'))   
        user = UserProfile.objects.get(pk=data.get('user')) 
        wb = Workbook()
        ws = wb.active
        ws.title = 'Purchase Report By User'
        ws.merge_cells('A2:K2')
        ws.merge_cells('A3:K3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Purchase Report By User'
        ws['A3'] = f'From {date_from} to {date_to}'
        ws['A4'] = ''
        ws.append(['DATE','STORE','INVOICE No.','USER','CUSTOMER','TOTAL','PROMOTION CODE','DISCOUNT(%)','NET TOTAL','PAYMENT METHOD','STATUS'])    
        sale_transactions = Purcase.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(user=user.user).order_by('invoice_number')
        for sale in sale_transactions:
            ws.append([sale['transaction_date'].strftime('%Y-%m-%d %H:%M:%S'), sale['store__name'],sale['invoice_number'],sale['user__username'],sale['full_name'].capitalize(), sale['total_amount'],sale['promotion_code__name'],sale['discount'],sale['net_amount'], sale['payment_method'], sale['status']])          
       
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')
def report_management_purchase_by_payment_method(request):
    form = PurchaseForm()
    form['payment_method'].initial = 'Cash'

    context = {
        'form' : form
    }

    return render(request, 'report_management_purchase_by_payment_method.html', context)

@in_groups('manager','superadmin')
def report_management_purchase_by_payment_method_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('start_date'))
        date_to = date.fromisoformat(data.get('end_date'))
        payment_method = data.get('payment_method')
        store = []
        for item in Store.objects.all().filter(is_selected = True):
            store.append(item.name)     
        sale_transactions = Purcase.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(payment_method=payment_method).order_by('invoice_number')
        if sale_transactions:
            sub_total = sale_transactions.values('store','store__name').annotate(total_sale=Sum('total_amount'), total_discount= Sum('discount_amount'),total_net_sale = Sum('net_amount'), total_promotion_code = Sum('promotion_code__amount_discount')).order_by('store')
            items={
                'sale_transactions' : list(sale_transactions),
                'sub_total' : list(sub_total),
                'store' : store
            }
        else:
            items={
                'message' : 'Record not found!'
            }
        return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def report_management_purchase_by_payment_method_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_to = date.fromisoformat(data.get('end_date'))
        date_from = date.fromisoformat(data.get('start_date'))   
        payment_method = data.get('payment_method') 
        wb = Workbook()
        ws = wb.active
        ws.title = 'Purchase Report By Payment Method'
        ws.merge_cells('A2:K2')
        ws.merge_cells('A3:K3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Sale Report By Invoice'
        ws['A3'] = f'From {date_from} to {date_to}'
        ws['A4'] = ''
        ws.append(['DATE','STORE','INVOICE No.','USER','CUSTOMER','TOTAL','PROMOTION CODE','DISCOUNT(%)','NET TOTAL','PAYMENT METHOD','STATUS'])    
        sale_transactions = Purcase.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(payment_method=payment_method).order_by('invoice_number')
        for sale in sale_transactions:
            ws.append([sale['transaction_date'].strftime('%Y-%m-%d %H:%M:%S'), sale['store__name'],sale['invoice_number'],sale['user__username'],sale['full_name'].capitalize(), sale['total_amount'],sale['promotion_code__name'],sale['discount'],sale['net_amount'], sale['payment_method'], sale['status']])          
       
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')    
def report_management_purchase_by_status(request):
    form = PurchaseForm()
    form['status'].initial = 'Paid'

    context = {
        'form' : form
    }

    return render(request, 'report_management_purchase_by_status.html', context)

@in_groups('manager','superadmin')
def report_management_purchase_by_status_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('start_date'))
        date_to = date.fromisoformat(data.get('end_date'))
        status = data.get('status')
        store = []
        for item in Store.objects.all().filter(is_selected = True):
            store.append(item.name)   
        if status == "Deleted":
            sale_transactions = Purcase.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(transaction_date__date__range = (date_from,date_to)).filter(status=status).order_by('invoice_number')
        else:
            sale_transactions = Purcase.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(status=status).order_by('invoice_number')
        if sale_transactions:
            sub_total = sale_transactions.values('store','store__name').annotate(total_sale=Sum('total_amount'), total_discount= Sum('discount_amount'),total_net_sale = Sum('net_amount'), total_promotion_code = Sum('promotion_code__amount_discount')).order_by('store')
            items={
                'sale_transactions' : list(sale_transactions),
                'sub_total' : list(sub_total),
                'store' : store
            }
        else:
            items={
                'message' : 'Record not found!'
            }

    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def report_management_purchase_by_status_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_to = date.fromisoformat(data.get('end_date'))
        date_from = date.fromisoformat(data.get('start_date'))   
        status = data.get('status') 
        wb = Workbook()
        ws = wb.active
        ws.title = 'Purchase Report By Stauts'
        ws.merge_cells('A2:K2')
        ws.merge_cells('A3:K3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Purchase Report By Stauts'
        ws['A3'] = f'From {date_from} to {date_to}'
        ws['A4'] = ''
        ws.append(['DATE','STORE','INVOICE No.','USER','CUSTOMER','TOTAL','PROMOTION CODE','DISCOUNT(%)','NET TOTAL','PAYMENT METHOD','STATUS'])    
        if status == "Deleted":
            sale_transactions = Purcase.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(transaction_date__date__range = (date_from,date_to)).filter(status=status).order_by('invoice_number')
        else:
            sale_transactions = Purcase.objects.values('store__name','invoice_number','transaction_date','user__username','customer','total_amount','discount','promotion_code__name','net_amount','payment_method','status','id').annotate(full_name=Concat('customer__first_name', Value(' '), 'customer__last_name'), discount_amount= (F('total_amount') * F('discount')/100)).filter(deleted_at__isnull=True).filter(transaction_date__date__range = (date_from,date_to)).filter(status=status).order_by('invoice_number')
        
        for sale in sale_transactions:
            ws.append([sale['transaction_date'].strftime('%Y-%m-%d %H:%M:%S'), sale['store__name'],sale['invoice_number'],sale['user__username'],sale['full_name'].capitalize(), sale['total_amount'],sale['promotion_code__name'],sale['discount'],sale['net_amount'], sale['payment_method'], sale['status']])          
       
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')    
def report_management_inventory_items(request):
    
    return render(request, 'report_management_inventory.html')

@in_groups('manager','superadmin')
def report_management_inventory_items_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        option = data.get('option')
        store = []
        for item in Store.objects.all().filter(is_selected = True):
            store.append(item.name) 
        if option == 'all':
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name','store').filter(status='Active').order_by('id').annotate(total_price=F('price')*F('quantity'))
            if item:
                sub_total = item.values('store','store__name').annotate(total= Sum('total_price'), total_item = Sum('quantity')).order_by('store')
                items = {
                    'item' : list(item),
                    'sub_total' : list(sub_total),
                    'store' : store
                    }
            else:
                items={
                        'message' : 'Record not found!',        
            }
        else:
            date_from = date.fromisoformat(data.get('start_date'))
            date_to = date.fromisoformat(data.get('end_date'))
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name','store').filter(status='Active').filter(update_on__date__range=(date_from,date_to)).order_by('id').annotate(total_price=F('price')*F('quantity'))
            if item:
                sub_total = item.values('store','store__name').annotate(total= Sum('total_price'), total_item = Sum('quantity')).order_by('store')
                items = {
                    'item' : list(item),
                    'sub_total' : list(sub_total),
                    'store' : store
                    }
            else:
                items={
                        'message' : 'Record not found!'
                }
        return JsonResponse(items, safe=False)   

@in_groups('manager','superadmin')    
def report_management_inventory_items_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        option = data.get('option')    
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventory Report By Item'
        ws.merge_cells('A2:J2')
        ws.merge_cells('A3:J3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Inventory Report By Item'
        ws['A3'] = ''
        ws['A4'] = ''
        ws.append(['ID','NAME','STORE','CATEGORY','SIZE','PRICE','QUANTITY','TOTAL','REODER ALERT','CREATED DATE'])    
        if option == 'all':
            ws['A3'] = f'All Period'
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name').filter(status='Active').order_by('id').annotate(total_price=F('price')*F('quantity'))
        else:
            date_to = date.fromisoformat(data.get('end_date'))
            date_from = date.fromisoformat(data.get('start_date'))
            ws['A3'] = f'From {date_from} to {date_to}'
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name').filter(status='Active').filter(update_on__date__range=(date_from,date_to)).order_by('id').annotate(total_price=F('price')*F('quantity'))

        for data in item:
                ws.append([data['id'], data['name'],data['store__name'],data['category__name'],data['size'], data['price'],data['quantity'],data['total_price'],data['reorder_alert'], data['update_on'].strftime('%Y-%m-%d %H:%M:%S')])          
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')
def report_management_inventory_category(request):
    form = InventoryForm()
    context = {
        'form': form
    }
    return render(request, 'report_management_inventory_category.html',context)

@in_groups('manager','superadmin')
def report_management_inventory_category_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        option = data.get('option')
        category = data.get('category')
        store = []
        for item in Store.objects.all().filter(is_selected = True):
            store.append(item.name) 
        if option == 'all':
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name','store').filter(status='Active').order_by('id').filter(category=category).annotate(total_price=F('price')*F('quantity'))
            if item:
                sub_total = item.values('store','store__name').annotate(total= Sum('total_price'), total_item = Sum('quantity')).order_by('store')
                items = {
                    'item' : list(item),
                    'sub_total' : list(sub_total),
                    'store' : store
                    }
            else:
                items={
                        'message' : 'Record not found!'
                }
        else:
            date_from = date.fromisoformat(data.get('start_date'))
            date_to = date.fromisoformat(data.get('end_date'))
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name','store').filter(status='Active').filter(update_on__date__range=(date_from,date_to)).filter(category=category).order_by('id').annotate(total_price=F('price')*F('quantity'))
            if item:
                sub_total = item.values('store','store__name').annotate(total= Sum('total_price'), total_item = Sum('quantity')).order_by('store')
                items = {
                    'item' : list(item),
                    'sub_total' : list(sub_total),
                    'store' : store
                    }
            else:
                items={
                        'message' : 'Record not found!'
                }
        
        return JsonResponse(items, safe=False)  

@in_groups('manager','superadmin')    
def report_management_inventory_category_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        option = data.get('option')
        category = data.get('category')    
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventory Report By Category'
        ws.merge_cells('A2:J2')
        ws.merge_cells('A3:J3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Inventory Report By Category'
        ws['A3'] = ''
        ws['A4'] = ''
        ws.append(['ID','NAME','STORE','CATEGORY','SIZE','PRICE','QUANTITY','TOTAL','REODER ALERT','CREATED DATE'])    
        if option == 'all':
            ws['A3'] = f'All Period'
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name').filter(status='Active').filter(category=category).order_by('id').annotate(total_price=F('price')*F('quantity'))
        else:
            date_to = date.fromisoformat(data.get('end_date'))
            date_from = date.fromisoformat(data.get('start_date'))
            ws['A3'] = f'From {date_from} to {date_to}'
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name').filter(status='Active').filter(category=category).filter(update_on__date__range=(date_from,date_to)).order_by('id').annotate(total_price=F('price')*F('quantity'))

        for data in item:
                ws.append([data['id'], data['name'],data['store__name'],data['category__name'],data['size'], data['price'],data['quantity'],data['total_price'],data['reorder_alert'], data['update_on'].strftime('%Y-%m-%d %H:%M:%S')])          
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')    
def report_management_inventory_low_item(request):
    items = Inventory.objects.all().filter(status='Active').filter(quantity__lt=F('reorder_alert')).annotate(total_price=F('price')*F('quantity')).order_by('-quantity').order_by('store')
    context = {
        'items' : items
    }
    return render(request, 'report_management_inventory_low_item.html', context)

@in_groups('manager','superadmin')
def report_management_inventory_low_item_excel(request):
    if request.method == 'POST':   
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventory Report By Low Item'
        ws.merge_cells('A2:J2')
        ws.merge_cells('A3:J3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Inventory Report By Low Item'
        ws['A3'] = ''
        ws['A4'] = ''
        ws.append(['ID','NAME','STORE','CATEGORY','SIZE','PRICE','QUANTITY','TOTAL','REODER ALERT','CREATED DATE'])    
        ws['A3'] = f'All Period'
        item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name').filter(quantity__lt=F('reorder_alert')).annotate(total_price=F('price')*F('quantity')).order_by('-quantity')
        for data in item:
                ws.append([data['id'], data['name'],data['store__name'],data['category__name'],data['size'], data['price'],data['quantity'],data['total_price'],data['reorder_alert'], data['update_on'].strftime('%Y-%m-%d %H:%M:%S')])          
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')    
def report_management_inventory_status(request):
    form = InventoryForm()
    context = {
        'form': form
    }
    return render(request, 'report_management_inventory_status.html', context)

@in_groups('manager','superadmin')
def report_management_inventory_status_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        option = data.get('option')
        status = data.get('status')
        store = []
        for item in Store.objects.all().filter(is_selected = True):
            store.append(item.name)
        if option == 'all':
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name','store').filter(status=status).order_by('id').annotate(total_price=F('price')*F('quantity'))
            if item:
                sub_total = item.values('store','store__name').annotate(total= Sum('total_price'), total_item = Sum('quantity')).order_by('store')
                items = {
                    'item' : list(item),
                    'sub_total' : list(sub_total),
                    'store' : store
                    }
            else:
                items={
                        'message' : 'Record not found!'
                }
        else:
            date_from = date.fromisoformat(data.get('start_date'))
            date_to = date.fromisoformat(data.get('end_date'))
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name','store').filter(status=status).filter(update_on__date__range=(date_from,date_to)).order_by('id').annotate(total_price=F('price')*F('quantity'))
            if item:
                sub_total = item.values('store','store__name').annotate(total= Sum('total_price'), total_item = Sum('quantity')).order_by('store')
                items = {
                    'item' : list(item),
                    'sub_total' : list(sub_total),
                    'store' : store
                    }
            else:
                items={
                        'message' : 'Record not found!'
                }
        
        return JsonResponse(items, safe=False)  

@in_groups('manager','superadmin')    
def report_management_inventory_status_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        option = data.get('option')
        status = data.get('status')    
        wb = Workbook()
        ws = wb.active
        ws.title = 'Inventory Report By Status'
        ws.merge_cells('A2:J2')
        ws.merge_cells('A3:J3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Inventory Report By Status'
        ws['A3'] = ''
        ws['A4'] = ''
        ws.append(['ID','NAME','STORE','CATEGORY','SIZE','PRICE','QUANTITY','TOTAL','REODER ALERT','CREATED DATE'])    
        if option == 'all':
            ws['A3'] = f'All Period'
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name').filter(status=status).order_by('id').annotate(total_price=F('price')*F('quantity'))
        else:
            date_to = date.fromisoformat(data.get('end_date'))
            date_from = date.fromisoformat(data.get('start_date'))
            ws['A3'] = f'From {date_from} to {date_to}'
            item = Inventory.objects.values('id','name','category__name','size','price','quantity','reorder_alert','update_on','store__name').filter(status=status).filter(update_on__date__range=(date_from,date_to)).order_by('id').annotate(total_price=F('price')*F('quantity'))

        for data in item:
                ws.append([data['id'], data['name'],data['store__name'],data['category__name'],data['size'], data['price'],data['quantity'],data['total_price'],data['reorder_alert'], data['update_on'].strftime('%Y-%m-%d %H:%M:%S')])          
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')
def report_management_customer(request):
    form = CustomerForm()
    context = {
        'form' : form
    }
    return render(request, 'report_management_customer.html', context)

@in_groups('manager','superadmin')
def report_management_customer_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        option = data.get('option')
        status = data.get('status')
        membership = data.get('membership')
        if option == 'all':
            customers = Customer.objects.order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status' ) 
            if customers:
                items = {
                    'item' : list(customers),
                    }
            else:
                items={
                        'message' : 'Record not found!',        
            }
        else:
            date_from = date.fromisoformat(data.get('start_date'))
            date_to = date.fromisoformat(data.get('end_date'))
            if membership == '':
                customers = Customer.objects.filter(status=status).filter(created_on__date__range=(date_from,date_to)).order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status') 
                if customers:
                    items = {
                        'item' : list(customers),
                        }
                else:
                    items={
                            'message' : 'Record not found!',        
                }
            else:
                customers = Customer.objects.filter(membership=membership,status=status).filter(created_on__date__range=(date_from,date_to)).order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status' ) 
                if customers:
                    items = {
                    'item' : list(customers),
                            }
                else:
                    items={
                            'message' : 'Record not found!'
                    }
        
        return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')    
def report_management_customer_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        option = data.get('option')
        status = data.get('status')
        membership = data.get('membership')   
        wb = Workbook()
        ws = wb.active
        ws.title = 'Customer Report'
        ws.merge_cells('A2:J2')
        ws.merge_cells('A3:J3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Customer Report'
        ws['A3'] = ''
        ws['A4'] = ''
        ws.append(['ID','STORE','FIRST NAME','LAST NAME','GENDER','PHONE NUMBER','E-MAIL','MEMBERSHIP','JOINT DATE','STATUS'])    
        if option == 'all':
            ws['A3'] = f'All Period'
            customers = Customer.objects.order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status' ) 
        else:
            if membership == '':
                date_from = date.fromisoformat(data.get('start_date'))
                date_to = date.fromisoformat(data.get('end_date'))
                ws['A3'] = f'From {date_from} to {date_to}'
                customers = Customer.objects.filter(status=status).order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status')     
            else:
                date_to = date.fromisoformat(data.get('end_date'))
                date_from = date.fromisoformat(data.get('start_date'))
                ws['A3'] = f'From {date_from} to {date_to}'
                customers = Customer.objects.filter(membership=membership,status=status).filter(created_on__date__range=(date_from,date_to)).order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status' ) 
            
        for data in customers:
            ws.append([data['id'], data['store__name'],data['first_name'],data['last_name'],data['gender'], data['phone_number1'],data['e_mail'],data['membership'],data['created_on'].strftime('%Y-%m-%d %H:%M:%S'), data['status']])          
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')
def report_management_supplier(request):
    form = SupplierForm()
    context = {
        'form' : form
    }
    return render(request, 'report_management_supplier.html', context)

@in_groups('manager','superadmin')
def report_management_supplier_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        option = data.get('option')
        status = data.get('status')
        membership = data.get('membership')
        if option == 'all':
            customers = Supplier.objects.order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status' ) 
            if customers:
                items = {
                    'item' : list(customers),
                    }
            else:
                items={
                        'message' : 'Record not found!',        
            }
        else:
            date_from = date.fromisoformat(data.get('start_date'))
            date_to = date.fromisoformat(data.get('end_date'))
            if membership == '':
                customers = Supplier.objects.filter(status=status).filter(created_on__date__range=(date_from,date_to)).order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status') 
                if customers:
                    items = {
                        'item' : list(customers),
                        }
                else:
                    items={
                            'message' : 'Record not found!',        
                }
            else:
                customers = Supplier.objects.filter(membership=membership,status=status).filter(created_on__date__range=(date_from,date_to)).order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status' ) 
                if customers:
                    items = {
                    'item' : list(customers),
                            }
                else:
                    items={
                            'message' : 'Record not found!'
                    }
        
        return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def report_management_supplier_excel(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        option = data.get('option')
        status = data.get('status')
        membership = data.get('membership')   
        wb = Workbook()
        ws = wb.active
        ws.title = 'Supplier Report'
        ws.merge_cells('A2:J2')
        ws.merge_cells('A3:J3')
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A3'].alignment = Alignment(horizontal='center')
        ws['A2'] = 'Supplier Report'
        ws['A3'] = ''
        ws['A4'] = ''
        ws.append(['ID','STORE','FIRST NAME','LAST NAME','GENDER','PHONE NUMBER','E-MAIL','MEMBERSHIP','JOINT DATE','STATUS'])    
        if option == 'all':
            ws['A3'] = f'All Period'
            customers = Supplier.objects.order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status' ) 
        else:
            if membership == '':
                date_from = date.fromisoformat(data.get('start_date'))
                date_to = date.fromisoformat(data.get('end_date'))
                ws['A3'] = f'From {date_from} to {date_to}'
                customers = Supplier.objects.filter(status=status).order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status')     
            else:
                date_to = date.fromisoformat(data.get('end_date'))
                date_from = date.fromisoformat(data.get('start_date'))
                ws['A3'] = f'From {date_from} to {date_to}'
                customers = Supplier.objects.filter(membership=membership,status=status).filter(created_on__date__range=(date_from,date_to)).order_by('id').values('id','store__name','first_name','last_name','gender','phone_number1','e_mail','membership','created_on','status' ) 
            
        for data in customers:
            ws.append([data['id'], data['store__name'],data['first_name'],data['last_name'],data['gender'], data['phone_number1'],data['e_mail'],data['membership'],data['created_on'].strftime('%Y-%m-%d %H:%M:%S'), data['status']])          
                    
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_len + 3
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="users.xlsx"'
        return response

@in_groups('manager','superadmin')
def report_management_financial(request):
    user = User.objects.get(username = request.user)
    store = Store.objects.filter(is_selected = True)
    if store.count() >1 :
        store = user.userprofile.store
    else:        
        store = Store.objects.get(is_selected = True).name
    
    context = {
        'store' : store
    }
    return render(request, 'report_management_financial.html', context)

@in_groups('manager','superadmin')
def report_management_financial_generate(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        date_from = date.fromisoformat(data.get('start_date'))
        date_to = date.fromisoformat(data.get('end_date'))
        total_sale = get_profit(date_from,date_to)
        profit_data = get_profit(date_from, date_to)
        if total_sale == 0:
            message = 'Record not found!'
            expanse = None
            total_expanse = None
            net_profit = None
        else:
            message = None
            expanse = ExpanseItem.objects.values('category__name').annotate(total=Sum('amount'))
            total_expanse = expanse.aggregate(expanse=Sum('total'))
            expanse = list(expanse)
            net_profit = profit_data['profit']
        items = {
            'total_sale' : total_sale,
            'message' : message,
            'expanse' : expanse,
            'total_expanse' : total_expanse,
            'net_profit' : net_profit
        }
       
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def user_management(request):
    users = User.objects.exclude(userprofile__status='Inactive').filter(userprofile__store__is_selected=True)
    user = User.objects.get(username = request.user)
    form = UserForm()
    form1 = UserProfileForm()
    form_update = UserUpdateForm()
    if request.method == 'POST':
        form = UserForm(request.POST)
        form1 = UserProfileForm(request.POST, request.FILES)
        if form.is_valid() and form1.is_valid():
            with transaction.atomic():
                try:
                    new_user = form.save(commit=False)
                    new_user.username = form.cleaned_data['username'].lower()
                    new_user.save()
                    user_profile = form1.save(commit=False)
                    user_profile.user = new_user
                    user_profile.save()
                    group = Group.objects.get(name= form1.cleaned_data['group'])
                    new_user.groups.add(group)

                    file_path = Path(settings.BASE_DIR / 'core/static/data')/'setting.json'
                    if file_path.exists():
                        with open(file_path, 'r') as file:
                            data = json.load(file)
                    else: 
                        data = {}
                    store = Store.objects.get(name = form1.cleaned_data['store'])
                    try:
                        data[str(form.cleaned_data['username'].lower())]['default_store'] = str(store.name)
                    except:
                        data[str(form.cleaned_data['username'].lower())] = ({
                                                    'username' : str(form.cleaned_data['username'].lower()),
                                                    'default_store' : str(store.name),
                                                            })
                    with open(file_path, 'w') as file:
                        json.dump(data, file, indent=4)
                    
                    messages.success(request, f'username {form.cleaned_data['username'].lower()} was sucessfully created')
                except Exception as e:
                    transaction.set_rollback(True)
                    messages.error(request, e)
                    return redirect('user_management')
                except IntegrityError as e:  
                    transaction.set_rollback(True)  
                    messages.error(request, e)
                    return redirect('user_management')        
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '            
        messages.error(request, error)
        return redirect ('user_management') 
    else:
        form = UserForm()
        form1 = UserProfileForm()
        form1.fields['store'].initial = user.userprofile.store
        form1.fields['status'].choices = [('Active','Active'),('Block','Block')]
        form_update = UserUpdateForm()
            
    context = {
        'form': form,
        'form1' : form1,
        'users' : users,
        'form_update' : form_update
    }

    return render(request, 'user_management.html', context)

@in_groups('manager','superadmin')
def user_management_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        if sortby == 'username' or sortby == 'first_name' or sortby == 'last_name':
            items = list(User.objects.values('userprofile__store__name','username','first_name','last_name','userprofile__gender','userprofile__date_of_brith','userprofile__phone_number','userprofile__group__name','userprofile__created_on','userprofile__status').exclude(userprofile__status='Inactive').filter(userprofile__store__is_selected=True).order_by(f'{sort_method}{sortby}'))
        else:
            items = list(User.objects.values('userprofile__store__name','username','first_name','last_name','userprofile__gender','userprofile__date_of_brith','userprofile__phone_number','userprofile__group__name','userprofile__created_on','userprofile__status').exclude(userprofile__status='Inactive').filter(userprofile__store__is_selected=True).order_by(f'{sort_method}userprofile__{sortby}'))
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def user_management_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        if sortby == 'username' or sortby == 'first_name' or sortby == 'last_name':
            items = list(User.objects.values('userprofile__store__name','username','first_name','last_name','userprofile__gender','userprofile__date_of_brith','userprofile__phone_number','userprofile__group__name','userprofile__created_on','userprofile__status').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(Q(full_name__icontains=search_text) | Q(username__icontains=search_text)).exclude(userprofile__status='Inactive').filter(userprofile__store__is_selected=True).order_by(f'{sort_method}{sortby}'))
        else:
            items = list(User.objects.values('userprofile__store__name','username','first_name','last_name','userprofile__gender','userprofile__date_of_brith','userprofile__phone_number','userprofile__group__name','userprofile__created_on','userprofile__status').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(Q(full_name__icontains=search_text) | Q(username__icontains=search_text)).exclude(userprofile__status='Inactive').filter(userprofile__store__is_selected=True).order_by(f'{sort_method}userprofile__{sortby}'))     
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def user_management_get_user(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        username = data.get('username')
        request.session['temp_product_id'] = username 
        user = list(User.objects.values('id','userprofile__store','username','first_name','last_name','userprofile__gender','userprofile__date_of_brith','userprofile__phone_number','userprofile__group','userprofile__created_on','userprofile__status','email','userprofile__address','userprofile__image_profile').filter(username=username))   
    return JsonResponse(user, safe=False)

@in_groups('manager','superadmin')
def user_management_modify_user(request, id):
    if request.method == 'POST':
        update_user = User.objects.get(id = id)
        update_user_profile = UserProfile.objects.get(user = update_user)
        form = UserUpdateForm(request.POST, instance=update_user)
        form1 = UserProfileForm(request.POST, request.FILES, instance=update_user_profile)
        if form.is_valid() and form1.is_valid():
            with transaction.atomic():
                try:
                    form.save()
                    user = form1.save(commit=False)
                    user.number_attempts = 0
                    user.save()
                    group = Group.objects.filter(name=form1.cleaned_data['group'])
                    update_user.groups.set(group)
                    return JsonResponse({"message_sucess": "User was updated sucessfully."})
                except Exception as e:
                    transaction.set_rollback(True)
                    return JsonResponse({"message_error": str(e)})
                except IntegrityError as e:
                    transaction.set_rollback(True)
                    return JsonResponse({"message_error": str(e)})
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '            
        messages.error(request, error)
        return redirect ('user_management')       
    return JsonResponse({"error": "Invalid request"})

@in_groups('manager','superadmin')
def user_management_deactivate_user(request, username):
    if request.method == 'POST':
        try:
            user = UserProfile.objects.get(user__username=username)
            user.delete_custom()
            message = f'User {username} was successfully deactivated.'
            return JsonResponse({'message_sucess': message}, safe=False)  
        except Exception as e:
            return JsonResponse({'message_error': str(e)}, safe=False) 
        except IntegrityError as e:
            return JsonResponse({'message_error': str(e)}, safe=False)   
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def user_management_reset_password_user(request, username):
    if request.method == 'POST':
        data = json.loads(request.body)
        new_password = data.get('new_password')
        try:
            user = User.objects.get(username=username)
            user.set_password(new_password)
            user.save()
            message = f'User password {username} was successfully reset.'
            return JsonResponse({'message_sucess': message}, safe=False)  
        except Exception as e:
            return JsonResponse({'message_error': str(e)}, safe=False) 
        except IntegrityError as e:
            return JsonResponse({'message_error': str(e)}, safe=False)   
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def user_management_deactivate_user_list(request):
    users = User.objects.filter(userprofile__status='Inactive').filter(userprofile__store__is_selected=True)
    user  = User.objects.get(username = request.user)
    form1 = UserProfileForm()
    form1.fields['store'].initial = user.userprofile.user
    form_update = UserUpdateForm()  
    context = {
        'users' : users,
        'form1' : form1,
        'form_update' : form_update
    }
    
    return render(request, 'user_management_deactivate.html', context)

@in_groups('manager','superadmin')
def user_management_deactivate_user_list_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        if sortby == 'username' or sortby == 'first_name' or sortby == 'last_name':
            items = list(User.objects.values('userprofile__store__name','username','first_name','last_name','userprofile__gender','userprofile__date_of_brith','userprofile__phone_number','userprofile__group__name','userprofile__created_on','userprofile__status').filter(userprofile__status='Inactive').filter(userprofile__store__is_selected=True).order_by(f'{sort_method}{sortby}'))
        else:
            items = list(User.objects.values('userprofile__store__name','username','first_name','last_name','userprofile__gender','userprofile__date_of_brith','userprofile__phone_number','userprofile__group__name','userprofile__created_on','userprofile__status').filter(userprofile__status='Inactive').filter(userprofile__store__is_selected=True).order_by(f'{sort_method}userprofile__{sortby}'))
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def user_management_deactivate_user_list_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        if sortby == 'username' or sortby == 'first_name' or sortby == 'last_name':
            items = list(User.objects.values('userprofile__store__name','username','first_name','last_name','userprofile__gender','userprofile__date_of_brith','userprofile__phone_number','userprofile__group__name','userprofile__created_on','userprofile__status').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(Q(full_name__icontains=search_text) | Q(username__icontains=search_text)).filter(userprofile__status='Inactive').filter(userprofile__store__is_selected=True).order_by(f'{sort_method}{sortby}'))
        else:
            items = list(User.objects.values('userprofile__store__name','username','first_name','last_name','userprofile__gender','userprofile__date_of_brith','userprofile__phone_number','userprofile__group__name','userprofile__created_on','userprofile__status').annotate(full_name=Concat('first_name', Value(' '), 'last_name')).filter(Q(full_name__icontains=search_text) | Q(username__icontains=search_text)).filter(userprofile__status='Inactive').filter(userprofile__store__is_selected=True).order_by(f'{sort_method}userprofile__{sortby}'))     
    return JsonResponse(items, safe=False)

@in_groups('manager','superadmin')
def user_management_activate_user(request, username):
    if request.method == 'POST':
        try:
            user = UserProfile.objects.get(user__username=username)      
            user.status = 'Active'
            user.save()
            message = f'User {username} was successfully activated.'
            return JsonResponse({'message_sucess': message}, safe=False)
        except Exception as e:
            return JsonResponse({'message_error': str(e)}, safe=False)
        except IntegrityError as e:
            return JsonResponse({'message_error': str(e)}, safe=False)
    return JsonResponse({"error": "Invalid request"}, status=400)

@login_not_required
def login(request):
    form = LoginForm()
    error_login = None
    if request.method == 'POST':
        form= LoginForm(request.POST)
        username = request.POST.get('username')
        password = request.POST.get('password')
        try:
            user_log = UserProfile.all_objects.get(user__username = username)
            if user_log.number_attempts >= 3 or user_log.status == 'Block':
                error_login = 'Your account has been locked out. Please contact system admistration.'
            else:
                user = authenticate(request, username=username, password=password)    
                if user is not None:
                    auth_login(request, user)
                    user = User.objects.get(username=request.user)
                    user_log.number_attempts = 0
                    user_log.save()
                    Store.objects.all().update(is_selected = False)
                    if user.userprofile.store.name == 'head office':
                        Store.objects.all().update(is_selected = True)
                        return redirect('dashborad')
                    else:
                        selected_store = Store.objects.get(name=user.userprofile.store)
                        selected_store.is_selected = True
                        selected_store.save()
                        return redirect('dashborad') 
                if user is None and username is not None:
                    user_log.number_attempts += 1
                    user_log.save()
                    if user_log.number_attempts >=3:
                        user_log.status = 'Block'
                        user_log.save()
                        error_login = f'Your account has been locked out. Please contact system admistration.'
                    else:
                        error_login = f'Username and password not matched. {3 - user_log.number_attempts} attemps left!'
        except Exception as e:
            error_login = f'User did not exist. Please contact system admistration.'
    context = {
        'form' : form,
        'message' : error_login
    }
    return render(request,'login.html', context)

def logout(request):
    if request.method == 'POST':
        logout(request)
    return redirect('login')

def setting(request):
    store_option = ['all']
    all_store = Store.objects.all()
    for item in all_store:
        if item.name != 'head office':
            store_option.append(item.name)
    
    form = CustomPasswordChangeForm(user = request.user)
    if request.method == 'POST':
        form = CustomPasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            try:
                user = form.save()
                update_session_auth_hash(request, user)
                userprofile = UserProfile.objects.get(user__username = request.user)
                userprofile.password_change_date = timezone.now()
                userprofile.save()
                messages.success(request, 'Your password was sucessfully changed.')
            except Exception as e:
                messages.error(request, str(e))
            except IntegrityError as e:
                messages.error(request, str(e))    
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '            
        messages.error(request, error) 
    else:
        form = CustomPasswordChangeForm(user = request.user)
    
    context = {
        'store_option' : store_option,
        'form' : form,
    }
    
    return render(request,'setting.html', context)

def setting_get_user_setting(request):
    if request.method == 'POST':
        user = str(request.user)
        file_path = Path(settings.BASE_DIR / 'core/static/data')/'setting.json'
        if file_path.exists():
            with open(file_path, 'r') as file:
                data = json.load(file)
        default_store = data[user]['default_store']
    return JsonResponse({'default_store': default_store}, safe=False)

def setting_set_default_store(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        store = data.get('store')
        cap_store = store.title()
        file_path = Path(settings.BASE_DIR / 'core/static/data')/'setting.json'
        if file_path.exists():
            with open(file_path, 'r') as file:
                data = json.load(file)
        else: 
            data = {}
        try:
            data[str(request.user)]['default_store'] = store
        except:
            data[str(request.user)] = ({
           'username' : str(request.user),
           'default_store' : store,
                })
        with open(file_path, 'w') as file:
            json.dump(data, file, indent=4)
        
        try:    
            if store.lower() != 'all':
                Store.objects.all().update(is_selected = False)
                selected_store = Store.objects.get(name=store)
                selected_store.is_selected = True
                selected_store.save()
                message = f' {cap_store} was successfully selected as default.'
                return JsonResponse({'message_sucess': message}, safe=False)
            else:
                Store.objects.all().update(is_selected = True)
                message = f' {store} was successfully selected as default.'
                return JsonResponse({'message_sucess': message}, safe=False)
        except Exception as e:
            message = str(e)
        except IntegrityError as e:
            message = str(e)
        
        return JsonResponse({'message_error': message}, safe=False)
    return JsonResponse({"message_error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def store_management(request):
    items_store = Store.objects.all()
    form = StoreForm(request.POST)
    if request.method == 'POST':
        form = StoreForm(request.POST, request.FILES)
        if form.is_valid():            
            with transaction.atomic():
                try:
                    form.save()
                    messages.success(request, 'Store was created sucessfully.')
                    return redirect ('store_management')
                except Exception as e:
                    transaction.set_rollback(True)
                    messages.error(request, str(e))
                except IntegrityError as e:
                    transaction.set_rollback(True)
                    messages.error(request, str(e)) 
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '            
        messages.error(request, error)
        return redirect ('store_management') 
    else:
        form = StoreForm()
    
    context = {
        'items' : items_store,
        'form' : form,
    }

    return render(request, 'store_management.html', context)

@in_groups('manager','superadmin')
def store_management_get(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        id = data.get('id')
        request.session['temp_product_id'] = id 
        item = list(Store.objects.values('id','name','code','phone','address','logo','sale_target').filter(id=id))   
    return JsonResponse(item, safe=False)

@in_groups('manager','superadmin')
def store_management_modify(request, id):
    if request.method == 'POST':
        update_store = Store.objects.get(id = id)
        form = StoreForm(request.POST, request.FILES, instance=update_store)
        if form.is_valid():
            try:
                form.save()
                return JsonResponse({"message_sucess": "Store was updated sucessfully."})
            except Exception as e:
                return JsonResponse({"message_error": str(e)})
            except IntegrityError as e:
                return JsonResponse({"message_error": str(e)})        
        error = ''
        for field, errors in form.errors.items():
            detail = (str(list(errors)).translate(str.maketrans('', '', string.punctuation)))
            error = error + f'({field} : {detail}) '
        return JsonResponse({"message_error": str(error)})       
    return JsonResponse({"error": "Invalid request"}, status=400)

@in_groups('manager','superadmin')
def store_management_sort(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''
        items = list(Store.objects.values('id','name','code','phone','address','logo','sale_target').order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False) 

@in_groups('manager','superadmin')
def store_management_search(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        search_text = data.get('search_text')
        sortby = data.get('sortby')
        if data.get('status') == 1:
            sort_method = '-'
        else:
            sort_method = ''       
        items = list(Store.objects.values('id','name','code','phone','address','logo','sale_target').filter(Q(name__icontains=search_text) | Q(id__icontains=search_text)).order_by(f'{sort_method}{sortby}'))
        
    return JsonResponse(items, safe=False)

